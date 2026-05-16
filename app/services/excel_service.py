# app/services/excel_service.py
"""
Service de lecture et parsing du fichier Excel modèle.
Isolé de Flask — peut être utilisé et testé indépendamment.
Accepte indifféremment un Path local ou des bytes (contenu R2).
"""
import re
from pathlib import Path
from io import BytesIO
import pandas as pd
import openpyxl


def _to_file_like(source):
    """Normalise une source Path ou bytes en objet file-like pour pandas/openpyxl."""
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"Fichier Excel introuvable : {path}")
        return path
    # bytes ou BytesIO
    if isinstance(source, bytes):
        return BytesIO(source)
    return source  # déjà file-like


def get_sheet_names(source) -> list[str]:
    """Retourne la liste des onglets du fichier Excel."""
    file_like = _to_file_like(source)
    wb = openpyxl.load_workbook(file_like, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_sheet_data(source, sheet_name: str) -> dict:
    """
    Lit un onglet Excel et retourne un dict structuré :
    {
        pk_column: str,
        pks: list,
        cote_columns: list[str],
        all_columns: list[str],
        data: list[dict]
    }
    Lève ValueError si la colonne PK est introuvable.
    """
    file_like = _to_file_like(source)
    df = pd.read_excel(file_like, sheet_name=sheet_name)

    # Nettoyage NaN / NaT
    df = df.replace({pd.NA: None, pd.NaT: None})
    df = df.where(pd.notnull(df), None)

    columns = df.columns.tolist()

    # Détection colonne PK
    pk_column = next(
        (col for col in columns
         if "PK" in str(col).upper() or "KILOMETRIQUE" in str(col).upper()),
        None,
    )
    if pk_column is None:
        raise ValueError("Colonne PK introuvable dans l'onglet (cherche 'PK' ou 'KILOMETRIQUE')")

    # PK non vides
    pks = [pk for pk in df[pk_column].tolist() if pk is not None and str(pk).strip() != ""]

    # Colonnes de côtes = toutes les colonnes numériques sauf PK
    cote_columns = [
        col for col in columns
        if col != pk_column and pd.api.types.is_numeric_dtype(df[col])
    ]

    # Sérialisation sûre (NaN / Inf → None)
    clean_records = []
    for record in df.to_dict(orient="records"):
        clean = {}
        for key, value in record.items():
            if value is None:
                clean[key] = None
            elif isinstance(value, float):
                if value != value or value in (float("inf"), float("-inf")):
                    clean[key] = None
                else:
                    clean[key] = value
            elif isinstance(value, int):
                clean[key] = value
            else:
                clean[key] = str(value)
        clean_records.append(clean)

    return {
        "pk_column": pk_column,
        "pks": pks,
        "cote_columns": cote_columns,
        "all_columns": columns,
        "data": clean_records,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  PARSING PARAMÉTRIQUE — parse_modele_config()
#
#  Identification STABLE par PRÉFIXES, jamais par noms exacts de colonnes.
#
#  Contrat de stabilité :
#    - Préfixes colonnes (AXE_, TER_, AG_, AD_, GEO_, XY_) → immuables
#    - Noms PK_debut / PK_fin                               → immuables
#    - Valeurs PK format r'^\d+\+\d{3}$'                   → immuable
#    - Noms complets de colonnes / d'onglets                → peuvent changer
# ═══════════════════════════════════════════════════════════════════════════════

_PK_RE = re.compile(r'^\d+\+\d{3}$')

# Préfixes de colonnes utilisées dans les calculs
_CALC_PREFIXES = frozenset({'AXE', 'TER', 'AG', 'AD', 'GEO', 'XY', 'ASG', 'ASD', 'ASS'})

# Noms de colonnes structurelles (jamais des paramètres de calcul)
_STRUCT_COLS = frozenset({
    'NUM', 'NOM_SECTION', 'NOM', 'TITRE', 'DESCRIPTION',
    'PK_DEBUT', 'PK_FIN', 'PK', 'KILOMETRIQUE',
})


# ── Helpers atomiques ─────────────────────────────────────────────────────────

def _pk_to_m(pk_str):
    """'1+250' → 1250 (mètres). Retourne None si format invalide."""
    try:
        a, b = str(pk_str).strip().split('+')
        return int(a) * 1000 + int(b)
    except Exception:
        return None


def _is_pk_val(val):
    """True si val ressemble à un PK topographique (ex: '1+250')."""
    if val is None:
        return False
    if isinstance(val, float) and val != val:
        return False
    return bool(_PK_RE.match(str(val).strip()))


def _safe_float(v):
    """Convertit en float ou None (gère NaN, Inf, None, str)."""
    if v is None:
        return None
    try:
        f = float(v)
        return None if (f != f or abs(f) == float('inf')) else f
    except (TypeError, ValueError):
        return None


def _safe_str(v):
    """Convertit en str non-vide ou None (gère NaN, None)."""
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    s = str(v).strip()
    return s if s not in ('', 'nan', 'None', 'NaT') else None


def _detect_pk_col(df):
    """
    Trouve la colonne PK dans un DataFrame.
    1) Par nom exact (PK, KILOMETRIQUE)
    2) Par valeurs : toutes les premières valeurs non-null matchent _PK_RE
    """
    for col in df.columns:
        if str(col).upper().strip() in ('PK', 'KILOMETRIQUE'):
            return col
    for col in df.columns:
        sample = [v for v in df[col].dropna().head(5) if v is not None]
        if len(sample) >= 1 and all(_is_pk_val(v) for v in sample):
            return col
    return None


# ── Readers spécialisés ───────────────────────────────────────────────────────

_CLEAN_ID_RE = re.compile(r'^[A-Za-z][A-Za-z0-9_]*$')


def _find_header_row(df_raw):
    """
    Détecte la ligne d'en-têtes dans un DataFrame lu avec header=None.

    Un nom de colonne valide est un identifiant propre (lettres/chiffres/_,
    sans espaces ni caractères spéciaux). Les bandeaux de titre ne passent
    pas ce filtre car ils contiennent des tirets, espaces et accents.

    Retourne l'index (0-based) de la ligne, ou None si non trouvé.
    """
    for idx, row in df_raw.iterrows():
        # Garder uniquement les valeurs qui ressemblent à un identifiant
        clean = [
            str(v).strip() for v in row
            if _safe_str(v) is not None
            and _CLEAN_ID_RE.match(str(v).strip())
        ]
        if not clean:
            continue
        if any(v.upper() in ('PK', 'PK_DEBUT', 'PK_FIN', 'KILOMETRIQUE') for v in clean):
            return idx
        if any(v.split('_')[0].upper() in _CALC_PREFIXES for v in clean):
            return idx
    return None


def _apply_header(df_raw, header_idx):
    """Reconstruit un DataFrame avec la ligne header_idx comme en-têtes."""
    headers = [
        str(v).strip() if _safe_str(v) is not None else f'_col{i}'
        for i, v in enumerate(df_raw.iloc[header_idx])
    ]
    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = headers
    return df.reset_index(drop=True)


def _read_gen_info(df):
    """
    Lit un onglet GEN_ (structure clé/valeur : col A = clé, col B = valeur).
    Filtre les lignes dont la clé commence par 'GEN_'.
    """
    result = {}
    for _, row in df.iterrows():
        k = _safe_str(row.iloc[0])
        if not k or not k.upper().startswith('GEN_'):
            continue
        v = row.iloc[1] if len(row) > 1 else None
        fv = _safe_float(v)
        result[k] = fv if fv is not None else _safe_str(v)
    return result


def _read_profil_long(df, pk_col):
    """
    Lit AXE_PROFIL_LONG → liste de {pk, pk_m, z_axe}.
    La colonne Z est identifiée par son préfixe AXE_ (pas par son nom exact).
    """
    z_col = next(
        (c for c in df.columns
         if c != pk_col and str(c).split('_')[0].upper() == 'AXE'),
        None,
    )
    if z_col is None:
        return []
    rows = []
    for _, row in df.iterrows():
        pk = _safe_str(row[pk_col])
        if not pk or not _is_pk_val(pk):
            continue
        pk_m = _pk_to_m(pk)
        z    = _safe_float(row[z_col])
        if pk_m is not None and z is not None:
            rows.append({'pk': pk, 'pk_m': pk_m, 'z_axe': z})
    return sorted(rows, key=lambda r: r['pk_m'])


def _read_sections(df):
    """
    Lit un onglet à sections (PK_debut requis, PK_fin optionnel).
    Retourne liste de {pk_debut_m, pk_fin_m, params: {col_name: valeur}}.

    params contient TOUTES les colonnes dont le préfixe est dans _CALC_PREFIXES
    (valeurs numériques ET chaînes — ex: AG_Ref_Mode = 'AXLE').
    """
    col_debut = next(
        (c for c in df.columns if str(c).strip().upper() == 'PK_DEBUT'), None)
    col_fin = next(
        (c for c in df.columns if str(c).strip().upper() == 'PK_FIN'), None)
    if col_debut is None:
        return []

    calc_cols = [
        c for c in df.columns
        if str(c).strip().upper() not in _STRUCT_COLS
        and str(c).split('_')[0].upper() in _CALC_PREFIXES
    ]

    rows = []
    for _, row in df.iterrows():
        v_debut = _safe_str(row[col_debut])
        if not v_debut:
            continue
        pk_debut_m = _pk_to_m(v_debut)
        if pk_debut_m is None:
            continue

        pk_fin_m = None
        if col_fin:
            v_fin = _safe_str(row[col_fin])
            if v_fin:
                pk_fin_m = _pk_to_m(v_fin)

        params = {}
        for col in calc_cols:
            fv = _safe_float(row[col])
            if fv is not None:
                params[str(col)] = fv
            else:
                sv = _safe_str(row[col])
                if sv:
                    params[str(col)] = sv

        rows.append({
            'pk_debut_m': pk_debut_m,
            'pk_fin_m':   pk_fin_m,
            'params':     params,
        })

    return sorted(rows, key=lambda r: r['pk_debut_m'])


def _read_ter_points(df):
    """
    Lit TER_PROFIL_TYPE : points de mesure par côté, triés par section puis ordre.
    Identification des colonnes par leur préfixe TER_ (pas par nom exact).
    """
    col_debut = next(
        (c for c in df.columns if str(c).strip().upper() == 'PK_DEBUT'), None)
    col_fin = next(
        (c for c in df.columns if str(c).strip().upper() == 'PK_FIN'), None)
    if col_debut is None:
        return []

    def _find(keyword):
        return next(
            (c for c in df.columns if keyword.upper() in str(c).upper()), None)

    col_cote  = _find('TER_COTE')   or _find('COTE')
    col_ordre = _find('TER_ORDRE')  or _find('ORDRE')
    col_label = _find('TER_LABEL')  or _find('LABEL')
    col_dist  = _find('DIST_AXE')
    col_pente = _find('PENTE')

    rows = []
    for _, row in df.iterrows():
        v_debut = _safe_str(row[col_debut])
        if not v_debut:
            continue
        pk_debut_m = _pk_to_m(v_debut)
        if pk_debut_m is None:
            continue

        pk_fin_m = None
        if col_fin:
            v_fin = _safe_str(row[col_fin])
            if v_fin:
                pk_fin_m = _pk_to_m(v_fin)

        ordre_raw = _safe_float(row[col_ordre]) if col_ordre else None

        rows.append({
            'pk_debut_m': pk_debut_m,
            'pk_fin_m':   pk_fin_m,
            'cote':       _safe_str(row[col_cote])  if col_cote  else None,
            'ordre':      int(ordre_raw) if ordre_raw is not None else None,
            'label':      _safe_str(row[col_label]) if col_label else None,
            'dist_axe_m': _safe_float(row[col_dist])  if col_dist  else None,
            'pente_pct':  _safe_float(row[col_pente]) if col_pente else None,
        })

    return sorted(rows, key=lambda r: (r['pk_debut_m'], r['ordre'] or 0))


def _read_longitudinal(df, pk_col):
    """
    Lit un onglet longitudinal (une ligne par PK) : ASS_LONG, GEO_COORDONNEES.
    Collecte toutes les colonnes avec un préfixe dans _CALC_PREFIXES.
    """
    calc_cols = [
        c for c in df.columns
        if c != pk_col
        and str(c).split('_')[0].upper() in _CALC_PREFIXES
    ]
    rows = []
    for _, row in df.iterrows():
        pk = _safe_str(row[pk_col])
        if not pk or not _is_pk_val(pk):
            continue
        pk_m = _pk_to_m(pk)
        if pk_m is None:
            continue
        entry = {'pk': pk, 'pk_m': pk_m}
        for col in calc_cols:
            fv = _safe_float(row[col])
            if fv is not None:
                entry[str(col)] = fv
        rows.append(entry)
    return sorted(rows, key=lambda r: r['pk_m'])


def _build_available_elements(config):
    """
    Déduit la liste des éléments calculables depuis la config parsée.
    Retourne {groupe: [{key, label, description}, ...]}.
    Le label = nom de colonne sans son premier composant de préfixe.
    """
    elements = {}

    # AXE : couches de chaussée (colonnes e_XX dans AXE_SECTIONS)
    axe_params = set()
    for s in config['sections'].get('AXE', []):
        axe_params.update(
            k for k in s['params']
            if str(k).split('_')[0].upper() == 'AXE' and isinstance(s['params'][k], float)
        )
    if axe_params and config['profil_long']:
        elements['AXE'] = [
            {'key': k,
             'label': '_'.join(str(k).split('_')[1:]).replace('_', ' '),
             'description': 'Couche de chaussée'}
            for k in sorted(axe_params)
        ]

    # TER : points de mesure (un élément par label unique)
    ter_seen = {}
    for pt in config['ter_points']:
        lbl = pt.get('label')
        if lbl and lbl not in ter_seen:
            ter_seen[lbl] = pt
    if ter_seen:
        elements['TER'] = [
            {'key': lbl,
             'label': lbl.replace('_', ' '),
             'description': f"Côté {d.get('cote', '?')}"}
            for lbl, d in ter_seen.items()
        ]

    # ASG / ASD : géométrie canaux (colonnes numériques avec prefix AG_ ou AD_)
    for grp, col_pfx in (('ASG', 'AG'), ('ASD', 'AD')):
        grp_params = set()
        for s in config['sections'].get(grp, []):
            grp_params.update(
                k for k in s['params']
                if str(k).split('_')[0].upper() == col_pfx
                and isinstance(s['params'][k], float)
            )
        if grp_params:
            elements[grp] = [
                {'key': k,
                 'label': '_'.join(str(k).split('_')[1:]).replace('_', ' '),
                 'description': 'Assainissement ' + ('Gauche' if grp == 'ASG' else 'Droit')}
                for k in sorted(grp_params)
            ]

    return elements


# ── Fonctions publiques de calcul (utilisées par calculation_service) ─────────

def active_section(sections, pk_m):
    """
    Retourne la section active pour pk_m parmi une liste de sections triées.
    Règle : dernière section dont pk_debut_m ≤ pk_m
            (et pk_fin_m est None ou pk_m < pk_fin_m).
    """
    active = None
    for s in sections:
        if s['pk_debut_m'] <= pk_m:
            if s['pk_fin_m'] is None or pk_m < s['pk_fin_m']:
                active = s
        else:
            break  # sections triées par pk_debut_m
    return active


def interp_z(profil_long, pk_m):
    """
    Interpolation linéaire de z_axe pour un pk_m donné.
    Extrapolation plate aux extrémités (pas d'extrapolation affine).
    """
    if not profil_long:
        return None
    if pk_m <= profil_long[0]['pk_m']:
        return profil_long[0]['z_axe']
    if pk_m >= profil_long[-1]['pk_m']:
        return profil_long[-1]['z_axe']
    for i in range(len(profil_long) - 1):
        p0, p1 = profil_long[i], profil_long[i + 1]
        if p0['pk_m'] <= pk_m <= p1['pk_m']:
            t = (pk_m - p0['pk_m']) / (p1['pk_m'] - p0['pk_m'])
            return round(p0['z_axe'] + t * (p1['z_axe'] - p0['z_axe']), 4)
    return None


# ── Point d'entrée principal ──────────────────────────────────────────────────

def parse_modele_config(source) -> dict:
    """
    Lit tout le fichier Excel modèle et retourne une config structurée.

    Identification STABLE par préfixes — jamais par noms exacts :
      - Préfixe d'onglet  : sheet_name.split('_')[0]  → groupe logique
      - Préfixe de colonne: col_name.split('_')[0]     → domaine de calcul
      - Colonne PK        : valeurs matchent r'^\\d+\\+\\d{3}$'
      - Sections          : colonnes PK_debut / PK_fin (convention stable)

    Paramètre
    ---------
    source : Path | str | bytes | BytesIO
        Fichier Excel modèle RECEPTA.

    Retourne
    --------
    dict avec les clés :
      gen              : {GEN_*: valeur}           — paramètres généraux
      profil_long      : [{pk, pk_m, z_axe}]       — profil en long axe
      sections         : {préfixe: [section]}       — sections par groupe
      ter_points       : [{pk_debut_m, ..., label}] — points terrassement
      ass_long         : [{pk, pk_m, col: val}]     — fil d'eau
      geo_coords       : [{pk, pk_m, col: val}]     — coordonnées
      imprevus_pks     : [pk_str]                   — PK hors-plan
      available_elements: {groupe: [{key, label}]}  — éléments calculables
      errors           : [str]                      — warnings non bloquants
    """
    # Normaliser la source pour qu'elle soit réutilisable plusieurs fois
    if isinstance(source, (str, Path)):
        _path = Path(source)
        def _fl():
            return _path
    elif isinstance(source, bytes):
        def _fl():
            return BytesIO(source)
    else:
        # BytesIO ou autre file-like → consommer en bytes
        raw = source.read() if hasattr(source, 'read') else bytes(source)
        def _fl():
            return BytesIO(raw)

    config = {
        'gen':                {},
        'profil_long':        [],
        'sections':           {},
        'ter_points':         [],
        'ass_long':           [],   # ASS_* générique (legacy)
        'ass_long_g':         [],   # AG_* — côtes fil d'eau Gauche par PK
        'ass_long_d':         [],   # AD_* — côtes fil d'eau Droit  par PK
        'geo_coords':         [],
        'imprevus_pks':       [],
        'available_elements': {},
        'errors':             [],
    }

    # ── Noms des onglets ──────────────────────────────────────────────────────
    try:
        _wb = openpyxl.load_workbook(_fl(), read_only=True)
        sheet_names = _wb.sheetnames
        _wb.close()
    except Exception as e:
        config['errors'].append(f"Ouverture classeur : {e}")
        return config

    # ── Onglets GEN_ : lecture sans header (structure clé/valeur) ────────────
    gen_sheets = [s for s in sheet_names if s.split('_')[0].upper() == 'GEN']
    if gen_sheets:
        try:
            raw_gen = pd.read_excel(_fl(), sheet_name=gen_sheets, header=None)
            dfs_gen = raw_gen if isinstance(raw_gen, dict) else {gen_sheets[0]: raw_gen}
            for _, df in dfs_gen.items():
                config['gen'].update(_read_gen_info(df))
        except Exception as e:
            config['errors'].append(f"GEN_ : {e}")

    # ── Tous les autres onglets : lecture standard ────────────────────────────
    non_gen = [s for s in sheet_names if s.split('_')[0].upper() != 'GEN']
    if not non_gen:
        config['available_elements'] = _build_available_elements(config)
        return config

    try:
        all_dfs_raw = pd.read_excel(_fl(), sheet_name=non_gen, header=None)
    except Exception as e:
        config['errors'].append(f"Lecture onglets : {e}")
        config['available_elements'] = _build_available_elements(config)
        return config

    for sheet_name, df_raw in all_dfs_raw.items():
        sheet_pfx = sheet_name.split('_')[0].upper()
        sheet_up  = sheet_name.upper()

        try:
            # Détecter la ligne d'en-têtes (robuste : header=0 ou header=1 selon onglet)
            hdr_idx = _find_header_row(df_raw)
            if hdr_idx is None:
                continue
            df = _apply_header(df_raw, hdr_idx)

            pk_col  = _detect_pk_col(df)
            has_sec = any(str(c).strip().upper() == 'PK_DEBUT' for c in df.columns)

            if sheet_pfx == 'AXE':
                if has_sec:
                    config['sections']['AXE'] = _read_sections(df)
                elif pk_col:
                    config['profil_long'] = _read_profil_long(df, pk_col)

            elif sheet_pfx == 'TER':
                config['ter_points'] = _read_ter_points(df)

            elif sheet_pfx in ('ASG', 'ASD'):
                if has_sec:
                    config['sections'][sheet_pfx] = _read_sections(df)

            elif sheet_pfx == 'ASS':
                if pk_col and not has_sec:
                    config['ass_long'] = _read_longitudinal(df, pk_col)

            elif sheet_pfx == 'AG':
                # Onglet longitudinal côté Gauche (ex: AG_Z_fil_eau)
                if pk_col and not has_sec:
                    config['ass_long_g'] = _read_longitudinal(df, pk_col)

            elif sheet_pfx == 'AD':
                # Onglet longitudinal côté Droit (ex: AD_Z_fil_eau)
                if pk_col and not has_sec:
                    config['ass_long_d'] = _read_longitudinal(df, pk_col)

            elif sheet_pfx == 'GEO':
                if has_sec:
                    existing = config['sections'].get('GEO', [])
                    config['sections']['GEO'] = existing + _read_sections(df)
                elif pk_col:
                    config['geo_coords'] = _read_longitudinal(df, pk_col)

            elif 'IMPREVUS' in sheet_up:
                if pk_col:
                    config['imprevus_pks'] = [
                        str(row[pk_col])
                        for _, row in df.iterrows()
                        if _is_pk_val(row[pk_col])
                    ]
            # DIV_LEGENDE et autres → ignorés silencieusement

        except Exception as e:
            config['errors'].append(f"{sheet_name} : {e}")

    config['available_elements'] = _build_available_elements(config)
    return config
