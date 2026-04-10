#!/usr/bin/env python3
"""
add_coords_to_models.py
Ajoute 2 onglets de spatialisation aux 4 modèles Excel assainissement :
  - PK_Coordonnees     : X, Y, Z de l'axe + gisement + dist cumulée à chaque PK
  - OFFSETS_TRANSVERSAUX : distances transversales de chaque élément / axe
"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_DIR   = Path(__file__).parent.parent
MODELS_DIR = BASE_DIR / "data" / "modeles_recepta"

# ─── Couleurs ─────────────────────────────────────────────────────────────────
C_HEADER_COORD  = "1E3A5F"   # bleu foncé
C_HEADER_OFFSET = "0F4C35"   # vert foncé
C_AXE           = "00FFFF"   # cyan axe
C_GAUCHE        = "FEE2E2"   # fond rouge clair  côté G
C_DROIT         = "DBEAFE"   # fond bleu clair   côté D
C_NEUTRE        = "F1F5F9"   # gris clair
C_WHITE         = "FFFFFF"

# ─── Définition des 4 projets ─────────────────────────────────────────────────
PROJETS = [
    {
        "fichier"      : "P1_RN_2x2voies_TPC_3km.xlsx",
        "nom"          : "Route Nationale 2x2 voies TPC — 3 km",
        "x0"           : 839_420.000,
        "y0"           : 6_516_800.000,
        "z0"           : 105.850,
        "gisement"     : 42.5,       # direction NE
        "pk_step"      : 25,
        "dz_per_100m"  : 0.15,       # légère montée
        "offsets": [
            # (nom_element,  dist_axe_m,   col_cote_reference,    onglet_ref)
            ("AXE",               0.000, "—",               "—"),
            ("TPC_Gauche",       -1.500, "AXE_G",           "Cote_Gauche"),
            ("TPC_Droit",        +1.500, "AXE_D",           "Cote_Droit"),
            ("Bord_Chaussee_G",  -7.500, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Bord_Chaussee_D",  +7.500, "Tablier_Fini_D",  "Cote_Droit"),
            ("Canal_G (100x100)",-8.500, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Canal_D (60x60)",  +8.000, "Tablier_Fini_D",  "Cote_Droit"),
            ("Accotement_G",     -9.500, "Bord_Propriete_G","Cote_Gauche"),
            ("Accotement_D",     +9.000, "Bord_Propriete_D","Cote_Droit"),
            ("Bord_Propriete_G",-12.000, "Bord_Propriete_G","Cote_Gauche"),
            ("Bord_Propriete_D",+11.500, "Bord_Propriete_D","Cote_Droit"),
        ],
    },
    {
        "fichier"      : "P2_Urbain_Trottoirs_1.5km.xlsx",
        "nom"          : "Route Urbaine avec Trottoirs — 1.5 km",
        "x0"           : 841_750.000,
        "y0"           : 6_518_200.000,
        "z0"           : 98.320,
        "gisement"     : 355.0,      # direction N
        "pk_step"      : 25,
        "dz_per_100m"  : -0.08,      # légère descente
        "offsets": [
            ("AXE",               0.000, "—",               "—"),
            ("Bord_Chaussee_G",  -4.000, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Bord_Chaussee_D",  +4.000, "Tablier_Fini_D",  "Cote_Droit"),
            ("Trottoir_G",       -4.800, "Bord_Propriete_G","Cote_Gauche"),
            ("Trottoir_D",       +4.800, "Bord_Propriete_D","Cote_Droit"),
            ("Canal_G (80x80)",  -5.200, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Canal_D (40x40)",  +5.000, "Tablier_Fini_D",  "Cote_Droit"),
            ("Bord_Propriete_G", -7.000, "Bord_Propriete_G","Cote_Gauche"),
            ("Bord_Propriete_D", +7.000, "Bord_Propriete_D","Cote_Droit"),
        ],
    },
    {
        "fichier"      : "P3_Autoroute_2x3_5km.xlsx",
        "nom"          : "Autoroute 2x3 voies — 5 km",
        "x0"           : 836_100.000,
        "y0"           : 6_512_500.000,
        "z0"           : 112.450,
        "gisement"     : 18.0,       # direction NNE
        "pk_step"      : 25,
        "dz_per_100m"  : 0.25,
        "offsets": [
            ("AXE",                0.000, "—",               "—"),
            ("TPC_Gauche",        -2.000, "AXE_G",           "Cote_Gauche"),
            ("TPC_Droit",         +2.000, "AXE_D",           "Cote_Droit"),
            ("BAU_G",            -11.000, "Tablier_Fini_G",  "Cote_Gauche"),
            ("BAU_D",            +11.000, "Tablier_Fini_D",  "Cote_Droit"),
            ("Canal_G (120x120)",-12.500, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Canal_D (120x120)",+12.500, "Tablier_Fini_D",  "Cote_Droit"),
            ("Accotement_G",     -14.000, "Bord_Propriete_G","Cote_Gauche"),
            ("Accotement_D",     +14.000, "Bord_Propriete_D","Cote_Droit"),
            ("Bord_Propriete_G", -17.000, "Bord_Propriete_G","Cote_Gauche"),
            ("Bord_Propriete_D", +17.000, "Bord_Propriete_D","Cote_Droit"),
        ],
    },
    {
        "fichier"      : "P4_Rural_1x2_1km.xlsx",
        "nom"          : "Route Rurale 1x2 voies — 1 km",
        "x0"           : 833_650.000,
        "y0"           : 6_520_100.000,
        "z0"           : 142.780,
        "gisement"     : 285.0,      # direction WNW
        "pk_step"      : 25,
        "dz_per_100m"  : -0.35,
        "offsets": [
            ("AXE",               0.000, "—",               "—"),
            ("Bord_Chaussee_G",  -3.000, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Bord_Chaussee_D",  +3.000, "Tablier_Fini_D",  "Cote_Droit"),
            ("Canal_G (50x50)",  -3.700, "Tablier_Fini_G",  "Cote_Gauche"),
            ("Canal_D (50x50)",  +3.700, "Tablier_Fini_D",  "Cote_Droit"),
            ("Accotement_G",     -4.500, "Bord_Propriete_G","Cote_Gauche"),
            ("Accotement_D",     +4.500, "Bord_Propriete_D","Cote_Droit"),
            ("Bord_Propriete_G", -6.000, "Bord_Propriete_G","Cote_Gauche"),
            ("Bord_Propriete_D", +6.000, "Bord_Propriete_D","Cote_Droit"),
        ],
    },
]


# ─── Helpers style ─────────────────────────────────────────────────────────────

def hdr_font(color="FFFFFF", bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, font=None, fill_=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill_:  c.fill   = fill_
    if align:  c.alignment = align
    if border: c.border = border
    if num_fmt: c.number_format = num_fmt
    return c


# ─── Génération des coordonnées X, Y, Z ───────────────────────────────────────

def generate_coords(pk_labels, x0, y0, z0, gisement_deg, pk_step, dz_per_100m):
    """
    Génère des coordonnées Lambert 93 fictives mais cohérentes pour chaque PK.
    Légère courbure sinusoïdale pour simuler un tracé réel.
    """
    bearing_rad = math.radians(gisement_deg)
    rows = []
    for i, pk in enumerate(pk_labels):
        dist = i * pk_step
        # Variation angulaire douce (±3° sur 800m)
        angle_var = math.sin(dist / 800.0) * math.radians(3.0)
        b = bearing_rad + angle_var
        x   = round(x0 + dist * math.sin(b), 3)
        y   = round(y0 + dist * math.cos(b), 3)
        z   = round(z0 + (dist / 100.0) * dz_per_100m, 3)
        gis = round(math.degrees(b) % 360, 4)
        rows.append({
            "PK"          : pk,
            "X"           : x,
            "Y"           : y,
            "Z_axe"       : z,
            "Gisement_deg": gis,
            "Dist_cumul_m": dist,
        })
    return rows


# ─── Onglet PK_Coordonnees ─────────────────────────────────────────────────────

def write_pk_coords(wb, pk_labels, coords, nom_projet):
    # Supprimer si déjà existant
    if "PK_Coordonnees" in wb.sheetnames:
        del wb["PK_Coordonnees"]

    ws = wb.create_sheet("PK_Coordonnees")

    # Titre
    ws.merge_cells("A1:F1")
    set_cell(ws, 1, 1,
             f"COORDONNÉES DE L'AXE — {nom_projet}",
             font=Font(name="Calibri", bold=True, color="FFFFFF", size=12),
             fill_=fill(C_HEADER_COORD),
             align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    # Sous-titre
    ws.merge_cells("A2:F2")
    set_cell(ws, 2, 1,
             "Système de référence : Lambert 93 (EPSG:2154)  —  Données à renseigner avec les coordonnées réelles du levé",
             font=Font(name="Calibri", italic=True, color="475569", size=9),
             fill_=fill("F8FAFC"),
             align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 18

    # En-têtes colonnes
    headers = ["PK", "X (m)", "Y (m)", "Z axe (m)", "Gisement (°)", "Dist. cumulée (m)"]
    hints   = ["Point kilométrique", "Lambert 93 — Est", "Lambert 93 — Nord",
               "Cote NGF de l'axe chaussée", "Direction route (0°=N, 90°=E)", "Distance depuis l'origine"]
    col_widths = [12, 16, 16, 14, 16, 18]

    for ci, (h, hint, w) in enumerate(zip(headers, hints, col_widths), start=1):
        set_cell(ws, 3, ci, h,
                 font=hdr_font(),
                 fill_=fill(C_HEADER_COORD),
                 align=center(),
                 border=thin_border())
        set_cell(ws, 4, ci, hint,
                 font=Font(name="Calibri", italic=True, color="64748B", size=8),
                 fill_=fill("F1F5F9"),
                 align=Alignment(horizontal="center", vertical="center"),
                 border=thin_border())
        ws.column_dimensions[chr(64 + ci)].width = w

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 15

    # Données
    for ri, row in enumerate(coords, start=5):
        is_pk_entier = (row["Dist_cumul_m"] % 100 == 0)  # met en évidence les PK ronds
        bg = "EFF6FF" if is_pk_entier else C_WHITE

        vals = [row["PK"], row["X"], row["Y"], row["Z_axe"], row["Gisement_deg"], row["Dist_cumul_m"]]
        fmts = [None, "#,##0.000", "#,##0.000", "0.000", "0.0000", "0.0"]

        for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            f = Font(name="Calibri", bold=is_pk_entier, size=9,
                     color="1E3A5F" if is_pk_entier else "1E293B")
            set_cell(ws, ri, ci, v, font=f, fill_=fill(bg),
                     align=center(), border=thin_border(), num_fmt=fmt)

    # Freeze + filtre
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A3:F{4 + len(coords)}"

    # Note bas de page
    note_row = 5 + len(coords) + 1
    ws.merge_cells(f"A{note_row}:F{note_row}")
    set_cell(ws, note_row, 1,
             "IMPORTANT : Ces coordonnées sont générées automatiquement (projet démo). "
             "Remplacer par les coordonnées réelles issues du levé topographique ou du GPS.",
             font=Font(name="Calibri", italic=True, color="B45309", size=8),
             fill_=fill("FFFBEB"),
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[note_row].height = 30


# ─── Onglet OFFSETS_TRANSVERSAUX ──────────────────────────────────────────────

def write_offsets(wb, offsets, nom_projet):
    if "OFFSETS_TRANSVERSAUX" in wb.sheetnames:
        del wb["OFFSETS_TRANSVERSAUX"]

    ws = wb.create_sheet("OFFSETS_TRANSVERSAUX")

    # Titre
    ws.merge_cells("A1:E1")
    set_cell(ws, 1, 1,
             f"DISTANCES TRANSVERSALES / AXE — {nom_projet}",
             font=Font(name="Calibri", bold=True, color="FFFFFF", size=12),
             fill_=fill(C_HEADER_OFFSET),
             align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    # Schéma ASCII
    schema_lines = [
        "SECTION EN TRAVERS TYPE — Vue schématique (de gauche à droite)",
        "",
        "  ← GAUCHE (distances négatives)        AXE        DROITE (distances positives) →",
        "  ─────────────────────────────────────── 0 ────────────────────────────────────────",
        "  Bord_Prop_G ... Canal_G ... Bord_Chaussee_G   |   Bord_Chaussee_D ... Canal_D ... Bord_Prop_D",
    ]
    for i, line in enumerate(schema_lines):
        ws.merge_cells(f"A{2+i}:E{2+i}")
        set_cell(ws, 2+i, 1, line,
                 font=Font(name="Courier New", size=8,
                           color="0F4C35" if i == 0 else "475569",
                           bold=(i == 0)),
                 fill_=fill("F0FDF4"),
                 align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[2+i].height = 14

    hdr_row = 7
    ws.row_dimensions[hdr_row].height = 22

    # En-têtes
    headers = ["Élément", "Distance / axe (m)", "Côté", "Colonne cote référence", "Onglet source"]
    widths  = [28, 22, 10, 28, 20]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        set_cell(ws, hdr_row, ci, h,
                 font=hdr_font(),
                 fill_=fill(C_HEADER_OFFSET),
                 align=center(),
                 border=thin_border())
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Données
    for ri, (elem, dist, col_ref, onglet) in enumerate(offsets, start=hdr_row+1):
        if dist < 0:
            cote, bg = "Gauche", C_GAUCHE
        elif dist > 0:
            cote, bg = "Droit", C_DROIT
        else:
            cote, bg = "Axe", "E0FFFF"

        is_axe = (dist == 0.0)

        row_vals = [elem, dist, cote, col_ref, onglet]
        row_fmts = [None, "+0.000;-0.000;0.000", None, None, None]
        row_bold = [is_axe, is_axe, False, False, False]

        for ci, (v, fmt, bold) in enumerate(zip(row_vals, row_fmts, row_bold), start=1):
            set_cell(ws, ri, ci, v,
                     font=Font(name="Calibri", size=9, bold=bold,
                               color="0F766E" if is_axe else "1E293B"),
                     fill_=fill(bg),
                     align=center(),
                     border=thin_border(),
                     num_fmt=fmt)

    # Note d'utilisation
    note_row = hdr_row + len(offsets) + 2
    ws.merge_cells(f"A{note_row}:E{note_row+2}")
    note = (
        "UTILISATION :\n"
        "• La colonne 'Distance / axe' donne la position transversale de chaque élément (- = gauche, + = droite).\n"
        "• Combinée aux coordonnées XY de PK_Coordonnees, elle permet de calculer la position GPS de chaque élément "
        "via : X_elem = X_axe + dist × sin(gisement + 90°) / Y_elem = Y_axe + dist × cos(gisement + 90°)."
    )
    set_cell(ws, note_row, 1, note,
             font=Font(name="Calibri", italic=True, color="1E3A5F", size=8),
             fill_=fill("EFF6FF"),
             align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    for r in range(note_row, note_row+3):
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = f"A{hdr_row+1}"


# ─── Traitement principal ──────────────────────────────────────────────────────

def main():
    for p in PROJETS:
        path = MODELS_DIR / p["fichier"]
        if not path.exists():
            print(f"ABSENT  {p['fichier']}")
            continue

        wb = load_workbook(path)

        # Lire les PK depuis Cote_Gauche
        import pandas as pd
        df = pd.read_excel(path, sheet_name="Cote_Gauche")
        pk_col = next((c for c in df.columns if "PK" in str(c).upper()), None)
        if pk_col is None:
            print(f"ERREUR  {p['fichier']} — colonne PK introuvable")
            continue
        pk_labels = [str(v) for v in df[pk_col].dropna().tolist()]

        # Générer les coordonnées
        coords = generate_coords(
            pk_labels, p["x0"], p["y0"], p["z0"],
            p["gisement"], p["pk_step"], p["dz_per_100m"]
        )

        # Écrire les 2 nouveaux onglets
        write_pk_coords(wb, pk_labels, coords, p["nom"])
        write_offsets(wb, p["offsets"], p["nom"])

        wb.save(path)
        print(f"OK  {p['fichier']}  ({len(pk_labels)} PK, {len(p['offsets'])} offsets)")

    print("\nTermine.")


if __name__ == "__main__":
    main()
