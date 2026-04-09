# -*- coding: utf-8 -*-
"""
Génère 4 fichiers Excel modèles RECEPTA pour projets routiers variés.
Structure CORRECTE : ligne 1 = noms colonnes, ligne 2+ = données.
"""
import io, math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="888888")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def fill(hex_): return PatternFill("solid", fgColor=hex_)

def style_header_row(ws, col_groups, widths):
    """
    Met en forme la LIGNE 1 (vrais noms de colonnes) avec couleurs par groupe.
    col_groups = liste de (start_col, end_col, color_hex)
    widths     = liste de largeurs par colonne
    """
    for start, end, color in col_groups:
        for ci in range(start, end + 1):
            c = ws.cell(1, ci)
            c.fill = fill(color)
            c.font = Font(color="FFFFFF", bold=True, size=8)
            c.alignment = CENTER
            c.border = BORDER
    ws.row_dimensions[1].height = 38
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

def style_data(ws, nrows, ncols, fills_by_col):
    """Applique fond alternant sur les lignes de données."""
    for r in range(2, nrows + 2):
        alt = (r % 2 == 0)
        for ci in range(1, ncols + 1):
            c = ws.cell(r, ci)
            base = fills_by_col[ci - 1]
            # Légère alternance
            c.fill = fill(base) if not alt else fill(base)
            c.font = Font(size=9, bold=(ci == 1), color=("1E3A5F" if ci == 1 else "000000"))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r].height = 14

def gen_pks(pk_fin, pas):
    pks, pk_m = [], []
    pk = 0
    while pk <= pk_fin:
        pks.append(f"{pk//1000}+{pk%1000:03d}")
        pk_m.append(pk)
        pk += pas
    return pks, pk_m

# ═══════════════════════════════════════════════════════════════
#  PROJET 1 — Route nationale 2×2 voies avec TPC
#  BB6.25 — dévers G:2.5% D:2.5% — Caniveaux 100×100 et 80×80
# ═══════════════════════════════════════════════════════════════
def projet_1():
    Z0, PL, PK_FIN, PAS = 102.500, 0.012, 2000, 25
    DIST_VOIE = 3.50; DIST_BORD = 3.50*2 + 1.0  # 2 voies + BAU 1m
    PENTE_G = PENTE_D = 0.025
    EP_BB, EP_GB4, EP_GNT, EP_FOND = 0.06, 0.10, 0.15, 0.20

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        # Bord TPC (intérieur, à 0.50m de l'axe)
        z_tpc_g = round(z - 0.50 * PENTE_G, 3)
        z_tpc_d = round(z - 0.50 * PENTE_D, 3)
        # Voie 1 (intérieure) = à 3.50m
        z_v1g = round(z - 3.50 * PENTE_G, 3)
        z_v1d = round(z - 3.50 * PENTE_D, 3)
        # Voie 2 (extérieure) + BAU = à 7.0 + 1.0 = 8.0m
        z_bau_g = round(z - 8.0 * PENTE_G, 3)
        z_bau_d = round(z - 8.0 * PENTE_D, 3)
        # Bordure = bord chaussée
        z_brd_g = round(z_bau_g - 0.02, 3)
        z_brd_d = round(z_bau_d - 0.02, 3)
        # Couches sous bord ext
        def couches(zb):
            return (round(zb, 3), round(zb-EP_BB, 3),
                    round(zb-EP_BB-EP_GB4, 3),
                    round(zb-EP_BB-EP_GB4-EP_GNT, 3))
        g_bb,g_gb4,g_gnt,g_fond = couches(z_bau_g)
        d_bb,d_gb4,d_gnt,d_fond = couches(z_bau_d)
        # Caniveau G : 100×100, tablier = bord, radier = bord-0.15-1.00
        g_tabl = round(z_brd_g, 3)
        g_rad  = round(z_brd_g - 0.15 - 1.00, 3)
        g_sem  = round(g_rad - 0.10, 3)
        g_bp   = round(z_brd_g + 0.20, 3)
        # Caniveau D : 80×80, tablier = bord, radier = bord-0.12-0.80
        d_tabl = round(z_brd_d, 3)
        d_rad  = round(z_brd_d - 0.12 - 0.80, 3)
        d_sem  = round(d_rad - 0.10, 3)
        d_bp   = round(z_brd_d + 0.20, 3)

        rows_g.append({"PK": lb,
            "AXE_BB": z, "AXE_GB4": round(z-EP_BB,3),
            "AXE_GNT": round(z-EP_BB-EP_GB4,3), "AXE_Fond": round(z-EP_BB-EP_GB4-EP_GNT,3),
            "G_TPC_Bord": z_tpc_g, "G_Voie1_Bord": z_v1g,
            "G_BAU_Bord": z_bau_g,
            "G_BB_Roul": g_bb, "G_GB4_Base": g_gb4, "G_GNT_SBase": g_gnt, "G_Fond": g_fond,
            "G_Bordure": z_brd_g,
            "G_Tabl_100x100": g_tabl, "G_Rad_100x100": g_rad,
            "G_Semelle": g_sem, "G_BP": g_bp})
        rows_d.append({"PK": lb,
            "AXE_BB": z, "AXE_GB4": round(z-EP_BB,3),
            "AXE_GNT": round(z-EP_BB-EP_GB4,3), "AXE_Fond": round(z-EP_BB-EP_GB4-EP_GNT,3),
            "D_TPC_Bord": z_tpc_d, "D_Voie1_Bord": z_v1d,
            "D_BAU_Bord": z_bau_d,
            "D_BB_Roul": d_bb, "D_GB4_Base": d_gb4, "D_GNT_SBase": d_gnt, "D_Fond": d_fond,
            "D_Bordure": z_brd_d,
            "D_Tabl_80x80": d_tabl, "D_Rad_80x80": d_rad,
            "D_Semelle": d_sem, "D_BP": d_bp})
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre": "RN 2x2 voies avec TPC",
        "devers": "G:2.5% | D:2.5%",
        "largeur": "2x(2×3.50m voies + 1.0m BAU)",
        "couches": "BB6cm / GB4 10cm / GNT 15cm / FF 20cm",
        "assain_g": "Caniveau ferme 100x100 (H int=1.00m, dalle e=15cm)",
        "assain_d": "Caniveau ferme 80x80  (H int=0.80m, dalle e=12cm)",
        "pente_long": f"{PL*100:.1f}%",
        "z0": Z0,
        "pk_fin": PK_FIN,
        "pas": PAS,
        "profil_type": "2x2",
        "cv_g": "100x100", "cv_d": "80x80",
        "pente_g": 0.025, "pente_d": 0.025,
        "dist_g": 8.0, "dist_d": 8.0,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 2 — Route urbaine avec trottoirs et stationnement
#  Milieu urbain — dévers G:2% D:3% — Caniveaux avaloirs
# ═══════════════════════════════════════════════════════════════
def projet_2():
    Z0, PL, PK_FIN, PAS = 98.200, 0.008, 800, 25
    PENTE_G, PENTE_D = 0.020, 0.030
    EP_BB, EP_TRAV, EP_FOND = 0.06, 0.20, 0.25

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        # Côté G : chaussée 3.5m + stationnement 2.0m + trottoir 1.5m
        z_ch_g     = round(z - 3.50  * PENTE_G, 3)   # bord chaussée G
        z_stat_g   = round(z - 5.50  * PENTE_G, 3)   # bord stationnement
        z_trottG   = round(z - 7.00  * PENTE_G + 0.10, 3)  # trottoir surélevé +10cm
        z_caniv_g  = round(z_ch_g - 0.02, 3)         # fond caniveau = sous bordure
        z_bord_g   = round(z_ch_g - 0.02, 3)
        g_bb   = round(z_ch_g, 3)
        g_trav = round(z_ch_g - EP_BB, 3)
        g_fond = round(z_ch_g - EP_BB - EP_TRAV, 3)
        # Assain G : avaloir grille — radier = bord - 0.40m
        g_rad  = round(z_bord_g - 0.40, 3)
        g_bp   = round(z_trottG + 0.05, 3)

        # Côté D : chaussée 3.5m + trottoir 2.0m (pas de stationnement)
        z_ch_d     = round(z - 3.50  * PENTE_D, 3)
        z_trottD   = round(z - 5.50  * PENTE_D + 0.12, 3)
        z_bord_d   = round(z_ch_d - 0.02, 3)
        d_bb   = round(z_ch_d, 3)
        d_trav = round(z_ch_d - EP_BB, 3)
        d_fond = round(z_ch_d - EP_BB - EP_TRAV, 3)
        d_rad  = round(z_bord_d - 0.35, 3)
        d_bp   = round(z_trottD + 0.05, 3)

        rows_g.append({"PK": lb,
            "AXE_BB": z, "AXE_Trav": round(z-EP_BB,3), "AXE_Fond": round(z-EP_BB-EP_TRAV,3),
            "G_BB_Roul": g_bb, "G_Trav_Base": g_trav, "G_Fond_Forme": g_fond,
            "G_Bordure_Finie": z_bord_g,
            "G_Bord_Stationnement": z_stat_g,
            "G_Trottoir_Fini": z_trottG,
            "G_Radier_Avaloir": g_rad,
            "G_Bord_Propriete": g_bp})
        rows_d.append({"PK": lb,
            "AXE_BB": z, "AXE_Trav": round(z-EP_BB,3), "AXE_Fond": round(z-EP_BB-EP_TRAV,3),
            "D_BB_Roul": d_bb, "D_Trav_Base": d_trav, "D_Fond_Forme": d_fond,
            "D_Bordure_Finie": z_bord_d,
            "D_Trottoir_Fini": z_trottD,
            "D_Radier_Avaloir": d_rad,
            "D_Bord_Propriete": d_bp})
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre": "Route Urbaine avec Trottoirs et Stationnement",
        "devers": "G:2.0% | D:3.0%",
        "largeur": "G: chaussée 3.5m + stationnement 2.0m + trottoir 1.5m | D: chaussée 3.5m + trottoir 2.0m",
        "couches": "BB 6cm / Travaux 20cm / Fond de Forme 25cm",
        "assain_g": "Avaloir grille — radier 0.40m sous bordure",
        "assain_d": "Avaloir grille — radier 0.35m sous bordure",
        "pente_long": f"{PL*100:.1f}%",
        "z0": Z0, "pk_fin": PK_FIN, "pas": PAS,
        "profil_type": "urbain",
        "cv_g": "Avaloir G", "cv_d": "Avaloir D",
        "pente_g": 0.020, "pente_d": 0.030,
        "dist_g": 3.5, "dist_d": 3.5,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 3 — Autoroute 2×3 voies avec accotements et fossés
#  Dévers G:2.5% D:2.5% — Fossés trapézoïdaux — PK 0 à 5km
# ═══════════════════════════════════════════════════════════════
def projet_3():
    Z0, PL, PK_FIN, PAS = 115.000, 0.006, 5000, 25
    PENTE_G = PENTE_D = 0.025
    EP_BB, EP_GB4, EP_GNT, EP_FOND, EP_PST = 0.06, 0.12, 0.20, 0.30, 0.15

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        # 3 voies × 3.75m + BAU 3.0m + accotement 1.0m = 14.25m
        z_v3g  = round(z - 11.25 * PENTE_G, 3)  # bord voie 3
        z_bau_g = round(z - 14.25 * PENTE_G, 3) # bord BAU
        z_acc_g = round(z - 15.25 * PENTE_G, 3) # accotement
        z_brd_g = round(z_acc_g - 0.02, 3)
        # Couches sous accotement
        g_bb   = round(z_bau_g, 3)
        g_gb4  = round(z_bau_g - EP_BB, 3)
        g_gnt  = round(z_bau_g - EP_BB - EP_GB4, 3)
        g_fond = round(z_bau_g - EP_BB - EP_GB4 - EP_GNT, 3)
        g_pst  = round(g_fond - EP_PST, 3)   # couche PST
        # Fossé trapézoïdal G : fond 0.80m prof / berge 1.50m
        g_fond_fosse = round(z_brd_g - 0.80, 3)
        g_tete_fosse = round(z_brd_g + 0.10, 3)  # ras terrain naturel
        g_bp = round(z_brd_g + 0.30, 3)

        z_v3d  = round(z - 11.25 * PENTE_D, 3)
        z_bau_d = round(z - 14.25 * PENTE_D, 3)
        z_acc_d = round(z - 15.25 * PENTE_D, 3)
        z_brd_d = round(z_acc_d - 0.02, 3)
        d_bb   = round(z_bau_d, 3)
        d_gb4  = round(z_bau_d - EP_BB, 3)
        d_gnt  = round(z_bau_d - EP_BB - EP_GB4, 3)
        d_fond = round(z_bau_d - EP_BB - EP_GB4 - EP_GNT, 3)
        d_pst  = round(d_fond - EP_PST, 3)
        d_fond_fosse = round(z_brd_d - 0.80, 3)
        d_tete_fosse = round(z_brd_d + 0.10, 3)
        d_bp = round(z_brd_d + 0.30, 3)

        rows_g.append({"PK": lb,
            "AXE_BB": z, "AXE_GB4": round(z-EP_BB,3),
            "AXE_GNT": round(z-EP_BB-EP_GB4,3),
            "AXE_Fond": round(z-EP_BB-EP_GB4-EP_GNT,3),
            "AXE_PST":  round(z-EP_BB-EP_GB4-EP_GNT-EP_PST,3),
            "G_Bord_Voie3": z_v3g, "G_Bord_BAU": z_bau_g, "G_Accotement": z_acc_g,
            "G_BB_Roul": g_bb, "G_GB4_Base": g_gb4, "G_GNT_SBase": g_gnt,
            "G_Fond": g_fond, "G_PST": g_pst,
            "G_Fond_Fosse": g_fond_fosse, "G_Tete_Fosse": g_tete_fosse,
            "G_BP": g_bp})
        rows_d.append({"PK": lb,
            "AXE_BB": z, "AXE_GB4": round(z-EP_BB,3),
            "AXE_GNT": round(z-EP_BB-EP_GB4,3),
            "AXE_Fond": round(z-EP_BB-EP_GB4-EP_GNT,3),
            "AXE_PST":  round(z-EP_BB-EP_GB4-EP_GNT-EP_PST,3),
            "D_Bord_Voie3": z_v3d, "D_Bord_BAU": z_bau_d, "D_Accotement": z_acc_d,
            "D_BB_Roul": d_bb, "D_GB4_Base": d_gb4, "D_GNT_SBase": d_gnt,
            "D_Fond": d_fond, "D_PST": d_pst,
            "D_Fond_Fosse": d_fond_fosse, "D_Tete_Fosse": d_tete_fosse,
            "D_BP": d_bp})
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre": "Autoroute 2x3 voies avec BAU et Fosses",
        "devers": "G:2.5% | D:2.5%",
        "largeur": "3×3.75m voies + 3.0m BAU + 1.0m accotement",
        "couches": "BB 6cm / GB4 12cm / GNT 20cm / FF 30cm / PST 15cm",
        "assain_g": "Fosse trapezoidal G — fond prof. 0.80m",
        "assain_d": "Fosse trapezoidal D — fond prof. 0.80m",
        "pente_long": f"{PL*100:.1f}%",
        "z0": Z0, "pk_fin": PK_FIN, "pas": PAS,
        "profil_type": "autoroute",
        "cv_g": "Fosse G", "cv_d": "Fosse D",
        "pente_g": 0.025, "pente_d": 0.025,
        "dist_g": 15.25, "dist_d": 15.25,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 4 — Route secondaire rurale simple (RR)
#  1×2 voies — dévers G:3% D:4% — Fossés ouverts — PK 0 à 500m
# ═══════════════════════════════════════════════════════════════
def projet_4():
    Z0, PL, PK_FIN, PAS = 85.750, 0.020, 500, 25
    PENTE_G, PENTE_D = 0.030, 0.040
    EP_BB, EP_GNT, EP_FOND = 0.05, 0.20, 0.25

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        # 1 voie par sens 3.0m + accotement 1.0m
        z_brd_g = round(z - 4.0 * PENTE_G, 3)
        z_brd_d = round(z - 4.0 * PENTE_D, 3)
        # Couches G
        g_bb   = round(z_brd_g, 3)
        g_gnt  = round(z_brd_g - EP_BB, 3)
        g_fond = round(z_brd_g - EP_BB - EP_GNT, 3)
        # Fosse G ouvert triangulaire : fond = bord - 0.60m
        g_fond_f = round(z_brd_g - 0.60, 3)
        g_berge  = round(z_brd_g + 0.05, 3)
        g_bp     = round(z_brd_g + 0.50, 3)
        # Couches D
        d_bb   = round(z_brd_d, 3)
        d_gnt  = round(z_brd_d - EP_BB, 3)
        d_fond = round(z_brd_d - EP_BB - EP_GNT, 3)
        d_fond_f = round(z_brd_d - 0.50, 3)
        d_berge  = round(z_brd_d + 0.05, 3)
        d_bp     = round(z_brd_d + 0.50, 3)

        rows_g.append({"PK": lb,
            "AXE_BB": z, "AXE_GNT": round(z-EP_BB,3),
            "AXE_Fond": round(z-EP_BB-EP_GNT,3),
            "G_BB_Roul": g_bb, "G_GNT_Base": g_gnt, "G_Fond_Forme": g_fond,
            "G_Fond_Fosse": g_fond_f, "G_Berge_Fosse": g_berge,
            "G_Bord_Propriete": g_bp})
        rows_d.append({"PK": lb,
            "AXE_BB": z, "AXE_GNT": round(z-EP_BB,3),
            "AXE_Fond": round(z-EP_BB-EP_GNT,3),
            "D_BB_Roul": d_bb, "D_GNT_Base": d_gnt, "D_Fond_Forme": d_fond,
            "D_Fond_Fosse": d_fond_f, "D_Berge_Fosse": d_berge,
            "D_Bord_Propriete": d_bp})
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre": "Route Rurale Secondaire 1x2 voies",
        "devers": "G:3.0% | D:4.0%",
        "largeur": "2×3.0m voies + 1.0m accotement",
        "couches": "BB 5cm / GNT 20cm / Fond de Forme 25cm",
        "assain_g": "Fosse ouvert triangulaire G — fond 0.60m",
        "assain_d": "Fosse ouvert triangulaire D — fond 0.50m",
        "pente_long": f"{PL*100:.1f}%",
        "z0": Z0, "pk_fin": PK_FIN, "pas": PAS,
        "profil_type": "rural",
        "cv_g": "Fosse G", "cv_d": "Fosse D",
        "pente_g": 0.030, "pente_d": 0.040,
        "dist_g": 4.0, "dist_d": 4.0,
    }

# ═══════════════════════════════════════════════════════════════
#  DESSIN DU PROFIL EN TRAVERS (matplotlib)
# ═══════════════════════════════════════════════════════════════
def draw_profil(meta, output_path):
    fig, ax = plt.subplots(figsize=(14, 5))
    ax.set_facecolor("#F8FAFC")
    fig.patch.set_facecolor("#0D1F38")

    ptype  = meta["profil_type"]
    dg     = meta["dist_g"]
    dd     = meta["dist_d"]
    pg     = meta["pente_g"]
    pd_    = meta["pente_d"]
    z_axe  = 0.0  # centré à 0

    # Calcul des points clés
    z_bg   = -dg * pg
    z_bd   =  -dd * pd_

    # Couches (approximation visuelle)
    ep_total = 0.45
    layers_g = [
        (0.06, "#333333", "BB"),
        (0.10, "#666666", "GB4"),
        (0.15, "#999999", "GNT"),
        (0.20, "#C4A882", "Fond"),
    ]

    # Fond naturel
    ground_x = [-dg - 2.5, -dg, -dg + 0.01, dg - 0.01, dg, dg + 2.5]
    ground_y = [z_bg - 0.80, z_bg - 0.80, z_bg, z_bd, z_bd - 0.80, z_bd - 0.80]

    # Terrain naturel fond
    ax.fill_between([-dg-2.5, dg+2.5], [-2.5, -2.5], [0.5, 0.5],
                    color="#D4B896", alpha=0.3)

    # Axe → bord (surface chaussée)
    road_x = [-dg, 0, dg]
    road_y = [z_bg, z_axe, z_bd]

    # Dessiner les couches (gauche + droite)
    cum = 0
    coul_ep = [0.06, 0.10, 0.15, 0.20]
    coul_col = ["#2C2C2C", "#555555", "#888888", "#B8936A"]
    coul_lab = ["BB - Roulement (6cm)", "GB4 - Base (10cm)",
                "GNT - Sous-base (15cm)", "Fond de Forme (20cm)"]
    for ep, col, lab in zip(coul_ep, coul_col, coul_lab):
        top_g = z_bg - cum;       bot_g = z_bg - cum - ep
        top_d = z_bd - cum;       bot_d = z_bd - cum - ep
        top_0 = z_axe - cum;      bot_0 = z_axe - cum - ep
        px = [-dg, 0, dg, dg, 0, -dg]
        py = [top_g, top_0, top_d, bot_d, bot_0, bot_g]
        ax.fill(px, py, color=col, alpha=0.85, zorder=3)
        cum += ep

    # Surface route
    ax.plot([-dg, 0, dg], [z_bg, z_axe, z_bd], color="#00FFFF", lw=2.5, zorder=5)

    # Axe vertical
    ax.axvline(0, color="#FBBF24", lw=1.5, ls="--", zorder=6, alpha=0.8)
    ax.text(0, z_axe + 0.08, "AXE", ha="center", color="#FBBF24", fontsize=8, fontweight="bold")

    # Flèches dévers
    ax.annotate("", xy=(-dg, z_bg), xytext=(0, z_axe),
                arrowprops=dict(arrowstyle="->", color="#EF4444", lw=1.5))
    ax.text(-dg/2, (z_axe+z_bg)/2 + 0.12,
            f"Devers G {pg*100:.1f}%", color="#EF4444", fontsize=7.5, ha="center",
            fontweight="bold")
    ax.annotate("", xy=(dg, z_bd), xytext=(0, z_axe),
                arrowprops=dict(arrowstyle="->", color="#3B82F6", lw=1.5))
    ax.text(dg/2, (z_axe+z_bd)/2 + 0.12,
            f"Devers D {pd_*100:.1f}%", color="#3B82F6", fontsize=7.5, ha="center",
            fontweight="bold")

    # Côté GAUCHE — caniveau/fossé
    if "100x100" in meta["cv_g"] or "80x80" in meta["cv_g"]:
        # Caniveau fermé
        cx, cw, ch = -dg - 0.15, 1.15, 1.15
        rect = mpatches.FancyBboxPatch((cx, z_bg - ch), cw, ch,
                                        boxstyle="square,pad=0", ec="#9B59B6", fc="#E8D5F5", lw=2, zorder=4)
        ax.add_patch(rect)
        ax.text(cx + cw/2, z_bg - ch/2, meta["cv_g"], ha="center", va="center",
                color="#6B2D8B", fontsize=7, fontweight="bold")
    else:
        # Fossé ouvert
        fx = [-dg-0.2, -dg-1.0, -dg-1.8]
        fy = [z_bg, z_bg - 0.70, z_bg]
        ax.fill(fx, fy, color="#B8D4E8", alpha=0.7, zorder=4)
        ax.plot(fx, fy, color="#2563EB", lw=2, zorder=5)
        ax.text(-dg-1.0, z_bg-0.85, meta["cv_g"], ha="center", color="#1E40AF", fontsize=7)

    # Côté DROIT — caniveau/fossé
    if "60x60" in meta["cv_d"] or "80x80" in meta["cv_d"]:
        cx, cw, ch = dg, 0.84, 0.84
        rect = mpatches.FancyBboxPatch((cx, z_bd - ch), cw, ch,
                                        boxstyle="square,pad=0", ec="#E91E8C", fc="#FAE8F3", lw=2, zorder=4)
        ax.add_patch(rect)
        ax.text(cx + cw/2, z_bd - ch/2, meta["cv_d"], ha="center", va="center",
                color="#8B2D6B", fontsize=7, fontweight="bold")
    else:
        fx = [dg+0.2, dg+1.0, dg+1.8]
        fy = [z_bd, z_bd - 0.60, z_bd]
        ax.fill(fx, fy, color="#B8D4E8", alpha=0.7, zorder=4)
        ax.plot(fx, fy, color="#2563EB", lw=2, zorder=5)
        ax.text(dg+1.0, z_bd-0.75, meta["cv_d"], ha="center", color="#1E40AF", fontsize=7)

    # Cotes distances
    ax.annotate("", xy=(-dg, z_bg-1.1), xytext=(0, z_bg-1.1),
                arrowprops=dict(arrowstyle="<->", color="#94A3B8", lw=1.2))
    ax.text(-dg/2, z_bg-1.25, f"{dg:.2f} m", ha="center", color="#94A3B8", fontsize=7.5)
    ax.annotate("", xy=(dg, z_bd-1.1), xytext=(0, z_bd-1.1),
                arrowprops=dict(arrowstyle="<->", color="#94A3B8", lw=1.2))
    ax.text(dg/2, z_bd-1.25, f"{dd:.2f} m", ha="center", color="#94A3B8", fontsize=7.5)

    # Labels G / D
    ax.text(-dg - 0.5, z_bg + 0.20, "GAUCHE", color="#EF4444",
            fontsize=9, fontweight="bold", ha="center")
    ax.text(dg + 0.5, z_bd + 0.20, "DROIT", color="#3B82F6",
            fontsize=9, fontweight="bold", ha="center")

    # Légende couches
    patches = [mpatches.Patch(color=c, label=l)
               for c, l in zip(coul_col, coul_lab)]
    ax.legend(handles=patches, loc="lower center",
              bbox_to_anchor=(0.5, -0.02),
              ncol=4, fontsize=7, framealpha=0.9,
              facecolor="#1E3A5F", labelcolor="white",
              edgecolor="#00FFFF")

    # Titre
    ax.set_title(f"Profil en travers type — {meta['titre']}",
                 color="white", fontsize=10, fontweight="bold", pad=10)

    ax.set_xlim(-dg - 2.8, dd + 2.8)
    ax.set_ylim(min(z_bg, z_bd) - 1.5, z_axe + 0.5)
    ax.set_xlabel("Distance / axe (m)", color="#94A3B8", fontsize=8)
    ax.set_ylabel("Cote relative (m)", color="#94A3B8", fontsize=8)
    ax.tick_params(colors="#94A3B8")
    for spine in ax.spines.values():
        spine.set_edgecolor("#334155")
    ax.grid(True, alpha=0.2, color="#475569")

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close()
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════════════
#  CONSTRUCTION D'UN CLASSEUR
# ═══════════════════════════════════════════════════════════════
COLORS = {
    "pk":     ("1E3A5F", "EBF4FF"),
    "axe":    ("0F4C81", "E8F0F8"),
    "trav_g": ("1A6B3C", "EAF5EE"),
    "trav_d": ("155E2F", "E4F5E9"),
    "ass_g":  ("6B2D8B", "F3E8FA"),
    "ass_d":  ("8B2D6B", "FAE8F3"),
    "extra":  ("7B3F00", "F5ECD5"),
}

def build_workbook(df_g, df_m, meta, filename):
    wb = Workbook()

    def add_sheet(ws, df, side, col_groups_def, widths):
        # Écrire les données avec pandas (ligne 1 = vrais noms colonnes)
        for ci, col in enumerate(df.columns, 1):
            ws.cell(1, ci, col)
        for ri, row in df.iterrows():
            for ci, val in enumerate(row, 1):
                ws.cell(ri + 2, ci, val)

        # Colorier ligne 1 (headers)
        for ci, (col, (hdr_c, data_c)) in enumerate(
            zip(df.columns, [COLORS[g] for g in col_groups_def]), 1):
            c = ws.cell(1, ci)
            c.fill = fill(hdr_c)
            c.font = Font(color="FFFFFF", bold=True, size=8)
            c.alignment = CENTER
            c.border = BORDER

        ws.row_dimensions[1].height = 36

        # Colorier données
        for ri in range(2, len(df) + 2):
            for ci, (_, (_, data_c)) in enumerate(
                zip(df.columns, [COLORS[g] for g in col_groups_def]), 1):
                c = ws.cell(ri, ci)
                c.fill = fill(data_c)
                c.font = Font(size=9, bold=(ci == 1),
                              color="1E3A5F" if ci == 1 else "1A1A1A")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = BORDER
            ws.row_dimensions[ri].height = 14

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "B2"

    # ── CÔTÉ GAUCHE ───────────────────────────────────────────
    ws_g = wb.active
    ws_g.title = "Cote_Gauche"
    ng = len(df_g.columns)
    # Déterminer les groupes de couleurs par colonne
    cg_groups = []
    for col in df_g.columns:
        col_u = col.upper()
        if col_u == "PK":                    cg_groups.append("pk")
        elif col_u.startswith("AXE"):        cg_groups.append("axe")
        elif col_u.startswith("G_") and any(k in col_u for k in
            ["BB","GB4","GNT","FOND","TRAV","BAU","ACC","BORD_VOIE"]):
                                             cg_groups.append("trav_g")
        elif col_u.startswith("G_") and any(k in col_u for k in
            ["TABL","RAD","SEM","BP","FOSSE","BERGE","AVALOIR","TROTTOIR",
             "STAT","PROP","BORD_PROP"]):     cg_groups.append("ass_g")
        elif col_u.startswith("G_"):         cg_groups.append("extra")
        else:                                cg_groups.append("trav_g")
    widths_g = [10] + [11] * (ng - 1)
    add_sheet(ws_g, df_g, "Gauche", cg_groups, widths_g)

    # ── CÔTÉ DROIT ────────────────────────────────────────────
    ws_d = wb.create_sheet("Cote_Droit")
    nd = len(df_m.columns)
    cd_groups = []
    for col in df_m.columns:
        col_u = col.upper()
        if col_u == "PK":                    cd_groups.append("pk")
        elif col_u.startswith("AXE"):        cd_groups.append("axe")
        elif col_u.startswith("D_") and any(k in col_u for k in
            ["BB","GB4","GNT","FOND","TRAV","BAU","ACC","BORD_VOIE"]):
                                             cd_groups.append("trav_d")
        elif col_u.startswith("D_") and any(k in col_u for k in
            ["TABL","RAD","SEM","BP","FOSSE","BERGE","AVALOIR","TROTTOIR",
             "STAT","PROP","BORD_PROP"]):     cd_groups.append("ass_d")
        elif col_u.startswith("D_"):         cd_groups.append("extra")
        else:                                cd_groups.append("trav_d")
    widths_d = [10] + [11] * (nd - 1)
    add_sheet(ws_d, df_m, "Droit", cd_groups, widths_d)

    # ── LÉGENDE ───────────────────────────────────────────────
    ws_l = wb.create_sheet("LEGENDE")
    ws_l.column_dimensions["A"].width = 5
    ws_l.column_dimensions["B"].width = 68

    # Image profil
    img_buf = draw_profil(meta, None)
    img = XLImage(img_buf)
    img.width  = 920
    img.height = 330
    ws_l.add_image(img, "B2")
    # Espace pour l'image
    for r in range(2, 22):
        ws_l.row_dimensions[r].height = 16.5

    # Texte légende sous l'image
    specs = [
        (22, "SPECIFICATIONS DU PROJET", "0D1F38", "00FFFF", True, 12),
        (23, f"Titre         : {meta['titre']}", "1E3A5F", "FFFFFF", True, 10),
        (24, f"Devers        : {meta['devers']}", "1E3A5F", "FFFFFF", False, 10),
        (25, f"Largeurs      : {meta['largeur']}", "1E3A5F", "FFFFFF", False, 10),
        (26, f"Pente long.   : {meta['pente_long']}", "1E3A5F", "FFFFFF", False, 10),
        (27, f"Cote axe Z0   : {meta['z0']:.3f} m NGF", "1E3A5F", "FFFFFF", False, 10),
        (28, f"PK debut/fin  : 0+000 / {meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}", "1E3A5F", "FFFFFF", False, 10),
        (29, f"Pas PK        : {meta['pas']} m", "1E3A5F", "FFFFFF", False, 10),
        (30, "", None, None, False, 6),
        (31, "COUCHES DE CHAUSSEE", "0F4C81", "FFFFFF", True, 10),
        (32, f"  {meta['couches']}", "0F4C81", "FFFFFF", False, 10),
        (33, "", None, None, False, 6),
        (34, f"ASSAINISSEMENT GAUCHE — {meta['cv_g']}", "6B2D8B", "FFFFFF", True, 10),
        (35, f"  {meta['assain_g']}", "6B2D8B", "FFFFFF", False, 10),
        (36, "", None, None, False, 6),
        (37, f"ASSAINISSEMENT DROIT  — {meta['cv_d']}", "8B2D6B", "FFFFFF", True, 10),
        (38, f"  {meta['assain_d']}", "8B2D6B", "FFFFFF", False, 10),
        (39, "", None, None, False, 6),
        (40, "COLONNES COTE_GAUCHE", "1A6B3C", "FFFFFF", True, 10),
    ]
    for i, col in enumerate(df_g.columns):
        specs.append((41+i, f"  {col}", "1A6B3C", "FFFFFF", False, 9))
    n = len(df_g.columns)
    specs.append((41+n, "COLONNES COTE_DROIT", "155E2F", "FFFFFF", True, 10))
    for i, col in enumerate(df_m.columns):
        specs.append((42+n+i, f"  {col}", "155E2F", "FFFFFF", False, 9))

    for row, txt, bg, fg, bold, sz in specs:
        c = ws_l.cell(row, 2, txt)
        if bg:
            c.fill = fill(bg)
            c.font = Font(color=fg, bold=bold, size=sz)
        else:
            c.font = Font(size=sz or 9)
        c.alignment = Alignment(vertical="center", indent=1)
        ws_l.row_dimensions[row].height = sz + 5

    wb.save(filename)
    print(f"OK : {filename}  ({len(df_g)} PK x {len(df_g.columns)} cols)")

# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    projets = [
        (projet_1, "data/clients/RECEPTA_P1_RN2x2_TPC.xlsx"),
        (projet_2, "data/clients/RECEPTA_P2_Urbain_Trottoirs.xlsx"),
        (projet_3, "data/clients/RECEPTA_P3_Autoroute_2x3.xlsx"),
        (projet_4, "data/clients/RECEPTA_P4_Rural_Secondaire.xlsx"),
    ]
    for fn, path in projets:
        df_g, df_d, meta = fn()
        build_workbook(df_g, df_d, meta, path)

    print("\nVerification lecture RECEPTA :")
    import pandas as pd
    for _, path in projets:
        df = pd.read_excel(path, sheet_name="Cote_Gauche", nrows=3)
        pk_col = next((c for c in df.columns if "PK" in str(c).upper()), None)
        cote_cols = [c for c in df.columns
                     if c != pk_col and pd.api.types.is_numeric_dtype(df[c])]
        print(f"  {path.split('/')[-1]} -> PK='{pk_col}' | {len(cote_cols)} cotes : {cote_cols[:4]}...")
