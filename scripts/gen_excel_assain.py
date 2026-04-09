import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PARAMETRES
Z0_AXE     = 100.000
PENTE_LONG = 0.014
PAS_PK     = 25
PK_FIN     = 1000
PENTE_G    = 0.020
PENTE_D    = 0.035
DIST_G     = 4.50
DIST_D     = 4.50
EP_BB=0.06; EP_GB4=0.10; EP_GNT=0.15; EP_FOND=0.20

pks, pk_metres = [], []
pk = 0
while pk <= PK_FIN:
    pks.append(f"{pk//1000}+{pk%1000:03d}")
    pk_metres.append(pk)
    pk += PAS_PK

rows_g, rows_d = [], []
for pk_label, d in zip(pks, pk_metres):
    z = Z0_AXE + d * PENTE_LONG
    zBG = z - DIST_G * PENTE_G
    zBD = z - DIST_D * PENTE_D
    rows_g.append({
        "PK":               pk_label,
        "AXE_BB":           round(z, 3),
        "AXE_GB4":          round(z - EP_BB, 3),
        "AXE_GNT":          round(z - EP_BB - EP_GB4, 3),
        "AXE_Fondation":    round(z - EP_BB - EP_GB4 - EP_GNT, 3),
        "G_BB_Roulement":   round(zBG, 3),
        "G_GB4_Base":       round(zBG - EP_BB, 3),
        "G_GNT_SousBase":   round(zBG - EP_BB - EP_GB4, 3),
        "G_Fond_Forme":     round(zBG - EP_BB - EP_GB4 - EP_GNT, 3),
        "G_Bordure_Finie":  round(zBG - 0.02, 3),
        "G_Radier_Caniveau":round(zBG - 0.60, 3),
        "G_Tablier_Fini":   round(zBG - 0.60 + 0.12, 3),
        "G_Bord_Propriete": round(zBG - 0.60 + 0.12 + 0.10, 3),
    })
    rows_d.append({
        "PK":               pk_label,
        "AXE_BB":           round(z, 3),
        "AXE_GB4":          round(z - EP_BB, 3),
        "AXE_GNT":          round(z - EP_BB - EP_GB4, 3),
        "AXE_Fondation":    round(z - EP_BB - EP_GB4 - EP_GNT, 3),
        "D_BB_Roulement":   round(zBD, 3),
        "D_GB4_Base":       round(zBD - EP_BB, 3),
        "D_GNT_SousBase":   round(zBD - EP_BB - EP_GB4, 3),
        "D_Fond_Forme":     round(zBD - EP_BB - EP_GB4 - EP_GNT, 3),
        "D_Bordure_Finie":  round(zBD - 0.02, 3),
        "D_Radier_Caniveau":round(zBD - 0.70, 3),
        "D_Tablier_Fini":   round(zBD - 0.70 + 0.15, 3),
        "D_Bord_Propriete": round(zBD - 0.70 + 0.15 + 0.10, 3),
    })

df_g = pd.DataFrame(rows_g)
df_d = pd.DataFrame(rows_d)

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_sheet(ws, df, side_label):
    cols = df.columns.tolist()
    ncols = len(cols)

    # Ligne 1 : titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, f"RECEPTA - Cotes Theoriques BB 6.25 - COTE {side_label.upper()}")
    c.fill = PatternFill("solid", fgColor="0D1F38")
    c.font = Font(color="00FFFF", bold=True, size=11)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Ligne 2 : groupes
    groups = [
        (1, 1,  "PK",                        "1E3A5F"),
        (2, 5,  "AXE (commun G/D)",           "0F4C81"),
        (6, 9,  f"Chaussee {side_label} - couches", "1A6B3C"),
        (10,10, "Bordure",                    "7B3F00"),
        (11,13, f"Assainissement {side_label}","6B2D8B"),
    ]
    for c1, c2, label, color in groups:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, label)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(color="FFFFFF", bold=True, size=8)
        c.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Ligne 3 : sous-titres
    subtitles = [
        "PK",
        "BB\n(-0.00)", "GB4\n(-0.06)", "GNT\n(-0.16)", "Fond.Forme\n(-0.31)",
        "BB Roul.\n(-0.00)", "GB4 Base\n(-0.06)", "GNT SBase\n(-0.16)", "Fond.Forme\n(-0.31)",
        "Bordure\nFinie",
        "Radier\nCaniveau", "Tablier\nFini", "Bord\nPropriete",
    ]
    colors3 = [
        "1E3A5F",
        "0F4C81","0F4C81","0F4C81","0F4C81",
        "1A6B3C","1A6B3C","1A6B3C","1A6B3C",
        "7B3F00",
        "6B2D8B","6B2D8B","6B2D8B",
    ]
    for j, (txt, col) in enumerate(zip(subtitles, colors3), start=1):
        c = ws.cell(3, j, txt)
        c.fill = PatternFill("solid", fgColor=col)
        c.font = Font(color="FFFFFF", bold=True, size=8)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[3].height = 30

    # Données
    data_fills = [
        PatternFill("solid", fgColor="EBF4FF"),
        PatternFill("solid", fgColor="E8F0F8"),
        PatternFill("solid", fgColor="E8F0F8"),
        PatternFill("solid", fgColor="E8F0F8"),
        PatternFill("solid", fgColor="E8F0F8"),
        PatternFill("solid", fgColor="EAF5EE"),
        PatternFill("solid", fgColor="EAF5EE"),
        PatternFill("solid", fgColor="EAF5EE"),
        PatternFill("solid", fgColor="EAF5EE"),
        PatternFill("solid", fgColor="F5ECD5"),
        PatternFill("solid", fgColor="F3E8FA"),
        PatternFill("solid", fgColor="F3E8FA"),
        PatternFill("solid", fgColor="F3E8FA"),
    ]
    for r_idx, row in df.iterrows():
        er = r_idx + 4
        for ci, val in enumerate(row, start=1):
            c = ws.cell(er, ci, val)
            c.fill = data_fills[ci-1]
            c.font = Font(size=9, bold=(ci==1))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[er].height = 14

    widths = [10, 9,9,9,10, 10,9,10,10, 9, 10,10,12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B4"

wb = Workbook()
ws_g = wb.active
ws_g.title = "Cote_Gauche"
style_sheet(ws_g, df_g, "Gauche")

ws_d = wb.create_sheet("Cote_Droit")
style_sheet(ws_d, df_d, "Droit")

# Onglet legende
ws_l = wb.create_sheet("LEGENDE")
ws_l.column_dimensions["A"].width = 65
legend = [
    ("RECEPTA - Fichier cotes theoriques assainissement BB 6.25", "0D1F38", "00FFFF", True),
    ("", None, None, False),
    ("STRUCTURE PROFIL EN TRAVERS", "1E3A5F", "FFFFFF", True),
    ("AXE : cotes a l axe de la chaussee (communes G et D)", "0F4C81", "FFFFFF", False),
    ("BB = Beton Bitumineux - couche de roulement (e=6cm)", "0F4C81", "FFFFFF", False),
    ("GB4 = Grave-Bitume - couche de base (e=10cm)", "0F4C81", "FFFFFF", False),
    ("GNT = Grave Non Traitee - sous-base (e=15cm)", "0F4C81", "FFFFFF", False),
    ("Fond de Forme (e=20cm)", "0F4C81", "FFFFFF", False),
    ("", None, None, False),
    ("CHAUSSEE GAUCHE - devers 2% - dist axe/bord = 4.50m", "1A6B3C", "FFFFFF", True),
    ("CHAUSSEE DROITE - devers 3.5% - dist axe/bord = 4.50m", "1A6B3C", "FFFFFF", True),
    ("", None, None, False),
    ("ASSAINISSEMENT GAUCHE : caniveau trapezodal", "6B2D8B", "FFFFFF", True),
    ("  Radier = bord chaussee - 0.60m", "6B2D8B", "FFFFFF", False),
    ("  Tablier fini = Radier + 0.12m", "6B2D8B", "FFFFFF", False),
    ("  Bord Propriete = Tablier + 0.10m", "6B2D8B", "FFFFFF", False),
    ("", None, None, False),
    ("ASSAINISSEMENT DROIT : caniveau rectangulaire", "6B2D8B", "FFFFFF", True),
    ("  Radier = bord chaussee - 0.70m", "6B2D8B", "FFFFFF", False),
    ("  Tablier fini = Radier + 0.15m", "6B2D8B", "FFFFFF", False),
    ("  Bord Propriete = Tablier + 0.10m", "6B2D8B", "FFFFFF", False),
    ("", None, None, False),
    ("Pente longitudinale = 1.4%  |  Cote axe PK0+000 = 100.000m", "1E3A5F", "FFFFFF", False),
]
for r, (txt, bg, fg, bold) in enumerate(legend, start=1):
    c = ws_l.cell(r, 1, txt)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=bold, size=10)
    else:
        c.font = Font(size=10)
    c.alignment = Alignment(vertical="center")
    ws_l.row_dimensions[r].height = 18

output = "data/clients/modele_recepta_ASSAINISSEMENT.xlsx"
wb.save(output)
print(f"OK : {output}")
print(f"{len(df_g)} PK  |  PK0+000 -> PK1+000  |  pas 25m")
