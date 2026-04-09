"""
Génère le fichier Excel modèle RECEPTA — Assainissement + Terrassement
Côté Gauche : caniveau 100x100  |  Côté Droit : caniveau 60x60
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
#  PARAMÈTRES — modifier selon le projet réel
# ═══════════════════════════════════════════════════════════════
Z0_AXE      = 100.000   # Cote axe au PK 0+000 (m NGF)
PENTE_LONG  = 0.014     # Pente longitudinale (1.4%)
PAS_PK      = 25        # Intervalle entre PK (m)
PK_FIN      = 1000      # Dernier PK (m)

# Dévers transversaux
PENTE_G     = 0.020     # Dévers gauche  2.0%
PENTE_D     = 0.035     # Dévers droit   3.5%

# Demi-largeur chaussée axe → bord (m)
DIST_G      = 4.50
DIST_D      = 4.50

# Épaisseurs couches de chaussée (m)
EP_BB       = 0.06      # Béton Bitumineux (roulement)
EP_GB4      = 0.10      # Grave-Bitume (base)
EP_GNT      = 0.15      # Grave Non Traitée (sous-base)
EP_FOND     = 0.20      # Fond de Forme

# ── CANIVEAU GAUCHE : 100×100 ──────────────────────────────────
#    Section intérieure 1.00m × 1.00m
#    Dalle de couverture (tablier) : e = 0.15m
#    Mur : e = 0.15m (largeur extérieure = 1.30m)
CV_G_PROF   = 1.00      # Profondeur intérieure (hauteur utile)
CV_G_DALLE  = 0.15      # Épaisseur tablier/dalle
CV_G_MUR    = 0.15      # Épaisseur mur latéral
# Radier = niveau bord chaussée − profondeur − épaisseur dalle
# Radier intérieur fond = bord_chaussee − CV_G_PROF − CV_G_DALLE
# Tablier fini (dessus dalle) = bord_chaussee − CV_G_DALLE  ← légèrement sous le bord
# On positionne le caniveau juste derrière la bordure
CV_G_OFFSET = 0.10      # distance bordure → face extérieure caniveau

# ── CANIVEAU DROIT : 60×60 ────────────────────────────────────
#    Section intérieure 0.60m × 0.60m
#    Dalle de couverture : e = 0.12m
CV_D_PROF   = 0.60
CV_D_DALLE  = 0.12
CV_D_MUR    = 0.12
CV_D_OFFSET = 0.10

# Bord de propriété = au-delà du caniveau
BP_OFFSET   = 0.20      # m après face extérieure caniveau

# ═══════════════════════════════════════════════════════════════
#  GÉNÉRATION DES PK
# ═══════════════════════════════════════════════════════════════
pks, pk_m = [], []
pk = 0
while pk <= PK_FIN:
    pks.append(f"{pk // 1000}+{pk % 1000:03d}")
    pk_m.append(pk)
    pk += PAS_PK

# ═══════════════════════════════════════════════════════════════
#  CALCUL DES COTES
# ═══════════════════════════════════════════════════════════════
rows_g, rows_d = [], []

for label, d in zip(pks, pk_m):
    z = round(Z0_AXE + d * PENTE_LONG, 3)   # cote axe à ce PK

    # ── CÔTÉ GAUCHE ───────────────────────────────────────────
    zBG = round(z - DIST_G * PENTE_G, 3)    # cote bord chaussée G

    # Terrassement
    g_bb   = round(zBG, 3)
    g_gb4  = round(zBG - EP_BB, 3)
    g_gnt  = round(zBG - EP_BB - EP_GB4, 3)
    g_fond = round(zBG - EP_BB - EP_GB4 - EP_GNT, 3)
    g_bord = round(zBG - 0.02, 3)           # bordure 2cm sous roulement

    # Assainissement caniveau 100×100
    # Tablier fini = dessus de la dalle = bord chaussée (ras-voirie)
    g_tablier  = round(zBG, 3)              # tablier affleurant le bord chaussée
    g_radier   = round(g_tablier - CV_G_DALLE - CV_G_PROF, 3)  # fond intérieur
    g_semelle  = round(g_radier - 0.10, 3)  # semelle de fondation du caniveau
    g_bp       = round(zBG + BP_OFFSET, 3)  # bord de propriété (légèrement en remblai)

    # Axe (identique pour les deux côtés)
    ax_bb   = round(z, 3)
    ax_gb4  = round(z - EP_BB, 3)
    ax_gnt  = round(z - EP_BB - EP_GB4, 3)
    ax_fond = round(z - EP_BB - EP_GB4 - EP_GNT, 3)

    rows_g.append({
        "PK":                       label,
        # Axe
        "AXE_BB_Roulement":         ax_bb,
        "AXE_GB4_Base":             ax_gb4,
        "AXE_GNT_SousBase":         ax_gnt,
        "AXE_Fond_Forme":           ax_fond,
        # Terrassement gauche
        "G_BB_Roulement":           g_bb,
        "G_GB4_Base":               g_gb4,
        "G_GNT_SousBase":           g_gnt,
        "G_Fond_Forme":             g_fond,
        "G_Bordure_Finie":          g_bord,
        # Assainissement gauche 100x100
        "G_Tablier_Fini_100x100":   g_tablier,
        "G_Radier_Fond_100x100":    g_radier,
        "G_Semelle_Fondation":      g_semelle,
        "G_Bord_Propriete":         g_bp,
    })

    # ── CÔTÉ DROIT ────────────────────────────────────────────
    zBD = round(z - DIST_D * PENTE_D, 3)

    d_bb   = round(zBD, 3)
    d_gb4  = round(zBD - EP_BB, 3)
    d_gnt  = round(zBD - EP_BB - EP_GB4, 3)
    d_fond = round(zBD - EP_BB - EP_GB4 - EP_GNT, 3)
    d_bord = round(zBD - 0.02, 3)

    # Assainissement caniveau 60×60
    d_tablier = round(zBD, 3)
    d_radier  = round(d_tablier - CV_D_DALLE - CV_D_PROF, 3)
    d_semelle = round(d_radier - 0.10, 3)
    d_bp      = round(zBD + BP_OFFSET, 3)

    rows_d.append({
        "PK":                       label,
        # Axe (même que gauche)
        "AXE_BB_Roulement":         ax_bb,
        "AXE_GB4_Base":             ax_gb4,
        "AXE_GNT_SousBase":         ax_gnt,
        "AXE_Fond_Forme":           ax_fond,
        # Terrassement droit
        "D_BB_Roulement":           d_bb,
        "D_GB4_Base":               d_gb4,
        "D_GNT_SousBase":           d_gnt,
        "D_Fond_Forme":             d_fond,
        "D_Bordure_Finie":          d_bord,
        # Assainissement droit 60x60
        "D_Tablier_Fini_60x60":     d_tablier,
        "D_Radier_Fond_60x60":      d_radier,
        "D_Semelle_Fondation":      d_semelle,
        "D_Bord_Propriete":         d_bp,
    })

df_g = pd.DataFrame(rows_g)
df_d = pd.DataFrame(rows_d)

# ═══════════════════════════════════════════════════════════════
#  STYLES EXCEL
# ═══════════════════════════════════════════════════════════════
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

PALETTE = {
    "titre":    "0D1F38",
    "pk":       "1E3A5F",
    "axe":      "0F4C81",
    "trav_g":   "1A6B3C",
    "trav_d":   "155E2F",
    "assain_g": "6B2D8B",
    "assain_d": "8B2D6B",
    "data_pk":  "EBF4FF",
    "data_axe": "E8F0F8",
    "data_tg":  "EAF5EE",
    "data_td":  "E6F5EC",
    "data_ag":  "F3E8FA",
    "data_ad":  "FAE8F3",
}

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def style_sheet(ws, df, side_label, cv_label, trav_color, assain_color, data_trav, data_assain):
    ncols = len(df.columns)

    # Ligne 1 — titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, f"RECEPTA - Cotes Theoriques - COTE {side_label.upper()} | Caniveau {cv_label}")
    c.fill = make_fill(PALETTE["titre"])
    c.font = Font(color="00FFFF", bold=True, size=11)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Ligne 2 — groupes de colonnes
    groups = [
        (1,  1,  "PK",                         PALETTE["pk"]),
        (2,  5,  "AXE (commun G et D)",         PALETTE["axe"]),
        (6,  10, f"Terrassement {side_label}",  trav_color),
        (11, 14, f"Assainissement {side_label} {cv_label}", assain_color),
    ]
    for c1, c2, label, color in groups:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, label)
        c.fill = make_fill(color)
        c.font = Font(color="FFFFFF", bold=True, size=9)
        c.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Ligne 3 — sous-titres colonnes
    subtitles = [
        "PK",
        "BB\nRoul.\n(+0.00)", "GB4\nBase\n(-0.06)", "GNT\nS-Base\n(-0.16)", "Fond.\nForme\n(-0.31)",
        "BB\nRoul.", "GB4\nBase", "GNT\nS-Base", "Fond.\nForme", "Bordure\nFinie",
        "Tablier\nFini", "Radier\nFond", "Semelle\nFond.", "Bord\nPropriete",
    ]
    colors3 = [
        PALETTE["pk"],
        PALETTE["axe"], PALETTE["axe"], PALETTE["axe"], PALETTE["axe"],
        trav_color, trav_color, trav_color, trav_color, trav_color,
        assain_color, assain_color, assain_color, assain_color,
    ]
    for j, (txt, col) in enumerate(zip(subtitles, colors3), start=1):
        c = ws.cell(3, j, txt)
        c.fill = make_fill(col)
        c.font = Font(color="FFFFFF", bold=True, size=8)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[3].height = 36

    # Données
    fills_data = [
        make_fill(PALETTE["data_pk"]),
        make_fill(PALETTE["data_axe"]), make_fill(PALETTE["data_axe"]),
        make_fill(PALETTE["data_axe"]), make_fill(PALETTE["data_axe"]),
        make_fill(data_trav), make_fill(data_trav),
        make_fill(data_trav), make_fill(data_trav), make_fill(data_trav),
        make_fill(data_assain), make_fill(data_assain),
        make_fill(data_assain), make_fill(data_assain),
    ]
    for r_idx, row in df.iterrows():
        er = r_idx + 4
        for ci, val in enumerate(row, start=1):
            c = ws.cell(er, ci, val)
            c.fill = fills_data[ci - 1]
            c.font = Font(size=9, bold=(ci == 1))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            if ci == 1:
                c.font = Font(size=9, bold=True, color="1E3A5F")
        ws.row_dimensions[er].height = 14

    # Largeurs colonnes
    widths = [10, 10, 9, 10, 10,  10, 9, 10, 10, 9,  11, 11, 11, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════════════
#  CONSTRUCTION DU CLASSEUR
# ═══════════════════════════════════════════════════════════════
wb = Workbook()

ws_g = wb.active
ws_g.title = "Cote_Gauche"
style_sheet(ws_g, df_g, "Gauche", "100x100",
            PALETTE["trav_g"], PALETTE["assain_g"],
            PALETTE["data_tg"], PALETTE["data_ag"])

ws_d = wb.create_sheet("Cote_Droit")
style_sheet(ws_d, df_d, "Droit", "60x60",
            PALETTE["trav_d"], PALETTE["assain_d"],
            PALETTE["data_td"], PALETTE["data_ad"])

# ── ONGLET LÉGENDE ────────────────────────────────────────────
ws_l = wb.create_sheet("LEGENDE")
ws_l.column_dimensions["A"].width = 70
ws_l.column_dimensions["B"].width = 25

legend = [
    ("RECEPTA — Fichier Cotes Theoriques Assainissement + Terrassement", "0D1F38", "00FFFF", True, 14),
    ("", None, None, False, 8),
    ("STRUCTURE GENERALE DU PROFIL", "1E3A5F", "FFFFFF", True, 12),
    ("", None, None, False, 6),
    ("AXE — Colonnes communes aux deux cotes", "0F4C81", "FFFFFF", True, 11),
    ("  AXE_BB_Roulement  : Cote finie de roulement a l axe", "0F4C81", "FFFFFF", False, 11),
    ("  AXE_GB4_Base      : Cote sous couche BB (axe - 0.06m)", "0F4C81", "FFFFFF", False, 11),
    ("  AXE_GNT_SousBase  : Cote sous GB4 (axe - 0.16m)", "0F4C81", "FFFFFF", False, 11),
    ("  AXE_Fond_Forme    : Cote fond de forme (axe - 0.31m)", "0F4C81", "FFFFFF", False, 11),
    ("", None, None, False, 6),
    ("TERRASSEMENT GAUCHE — Devers 2% | Dist. axe/bord = 4.50m", "1A6B3C", "FFFFFF", True, 11),
    ("  G_BB_Roulement   : Cote BB bord gauche = axe - 4.50 x 2%", "1A6B3C", "FFFFFF", False, 11),
    ("  G_GB4_Base       : Cote sous BB", "1A6B3C", "FFFFFF", False, 11),
    ("  G_GNT_SousBase   : Cote sous GB4", "1A6B3C", "FFFFFF", False, 11),
    ("  G_Fond_Forme     : Cote fond de forme bord gauche", "1A6B3C", "FFFFFF", False, 11),
    ("  G_Bordure_Finie  : Cote finie de la bordure (bord - 0.02m)", "1A6B3C", "FFFFFF", False, 11),
    ("", None, None, False, 6),
    ("ASSAINISSEMENT GAUCHE — Caniveau 100x100 (section interieure 1.00m x 1.00m)", "6B2D8B", "FFFFFF", True, 11),
    ("  G_Tablier_Fini   : Dessus dalle = affleurant bord chaussee", "6B2D8B", "FFFFFF", False, 11),
    ("  G_Radier_Fond    : Fond interieur = Tablier - dalle(0.15m) - hauteur(1.00m)", "6B2D8B", "FFFFFF", False, 11),
    ("  G_Semelle_Fond.  : Semelle de fondation = Radier - 0.10m", "6B2D8B", "FFFFFF", False, 11),
    ("  G_Bord_Propriete : Limite parcelle = bord chaussee + 0.20m", "6B2D8B", "FFFFFF", False, 11),
    ("", None, None, False, 6),
    ("TERRASSEMENT DROIT — Devers 3.5% | Dist. axe/bord = 4.50m", "155E2F", "FFFFFF", True, 11),
    ("  D_BB_Roulement   : Cote BB bord droit = axe - 4.50 x 3.5%", "155E2F", "FFFFFF", False, 11),
    ("  (memes couches que gauche, decalees par la pente de 3.5%)", "155E2F", "FFFFFF", False, 11),
    ("", None, None, False, 6),
    ("ASSAINISSEMENT DROIT — Caniveau 60x60 (section interieure 0.60m x 0.60m)", "8B2D6B", "FFFFFF", True, 11),
    ("  D_Tablier_Fini   : Dessus dalle = affleurant bord chaussee", "8B2D6B", "FFFFFF", False, 11),
    ("  D_Radier_Fond    : Fond interieur = Tablier - dalle(0.12m) - hauteur(0.60m)", "8B2D6B", "FFFFFF", False, 11),
    ("  D_Semelle_Fond.  : Semelle de fondation = Radier - 0.10m", "8B2D6B", "FFFFFF", False, 11),
    ("  D_Bord_Propriete : Limite parcelle = bord chaussee + 0.20m", "8B2D6B", "FFFFFF", False, 11),
    ("", None, None, False, 6),
    ("PARAMETRES GENERAUX", "1E3A5F", "FFFFFF", True, 11),
    (f"  Cote axe PK0+000  = {Z0_AXE:.3f} m NGF", "1E3A5F", "FFFFFF", False, 11),
    (f"  Pente longitudinale = {PENTE_LONG*100:.1f}%", "1E3A5F", "FFFFFF", False, 11),
    (f"  Pas entre PK       = {PAS_PK} m", "1E3A5F", "FFFFFF", False, 11),
    (f"  PK debut / fin     = 0+000 / {PK_FIN//1000}+{PK_FIN%1000:03d}", "1E3A5F", "FFFFFF", False, 11),
]

for r, (txt, bg, fg, bold, sz) in enumerate(legend, start=1):
    c = ws_l.cell(r, 1, txt)
    if bg:
        c.fill = make_fill(bg)
        c.font = Font(color=fg, bold=bold, size=sz)
    else:
        c.font = Font(size=sz or 10)
    c.alignment = Alignment(vertical="center", indent=1)
    ws_l.row_dimensions[r].height = sz + 6 if sz else 8

# ═══════════════════════════════════════════════════════════════
#  SAUVEGARDE
# ═══════════════════════════════════════════════════════════════
output = "data/clients/modele_recepta_ASSAINISSEMENT_v2.xlsx"
wb.save(output)

print(f"Fichier : {output}")
print(f"{len(df_g)} PK  |  PK0+000 -> PK{PK_FIN//1000}+{PK_FIN%1000:03d}  |  pas {PAS_PK}m")
print()
print("Colonnes Cote_Gauche :", list(df_g.columns))
print()
print("Colonnes Cote_Droit  :", list(df_d.columns))
print()
print("Apercu Cote_Gauche (3 PK) :")
print(df_g.head(3).to_string(index=False))
print()
print("Apercu Cote_Droit (3 PK) :")
print(df_d.head(3).to_string(index=False))
