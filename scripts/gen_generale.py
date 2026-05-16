# -*- coding: utf-8 -*-
"""
RECEPTA — Génération du fichier Excel modèle TYPE complet.

11 onglets organisés par groupe, avec préfixes, couleurs de tabs,
épaisseurs signées et dropdowns PK sur toutes les colonnes PK_debut/PK_fin.

Groupes :
  GEN_  gris       Identité projet
  AXE_  navy       Profil en long + épaisseurs (base commune)
  TER_  vert       Terrassement — points de mesure
  ASG_  violet     Assainissement Gauche
  ASD_  bordeaux   Assainissement Droit
  ASS_  violet     Assainissement — profil fil d'eau (optionnel)
  GEO_  marron     Cartographie
  DIV_  dark       Divers (IMPREVUS, LEGENDE)

Utilisation :
    python scripts/gen_generale.py
Sortie :
    data/clients/RECEPTA_MODELE_TYPE.xlsx
"""
import io, os, sys, math
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from gen_modeles_v3 import (
    projet_1,
    build_imprevus_sheet,
    build_profil_long_sheet,
    draw_profil,
)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUT_PATH   = "data/clients/RECEPTA_MODELE_TYPE.xlsx"
OUT_PROJET = "data/clients/RECEPTA_P1_RN_2x2voies_3km.xlsx"
os.makedirs("data/clients", exist_ok=True)

# ── Palette des groupes ───────────────────────────────────────────────────────
GROUPS = {
    "GEN": {"tab": "374151", "hdr": "1F2937", "dat": "F9FAFB", "bdr": "9CA3AF"},
    "AXE": {"tab": "1E3A5F", "hdr": "0F4C81", "dat": "EBF4FF", "bdr": "93C5FD"},
    "TER": {"tab": "166534", "hdr": "15803D", "dat": "DCFCE7", "bdr": "86EFAC"},
    "ASG": {"tab": "6B21A8", "hdr": "5B21B6", "dat": "EDE9FE", "bdr": "A78BFA"},
    "ASD": {"tab": "9D174D", "hdr": "831843", "dat": "FCE7F3", "bdr": "F9A8D4"},
    "ASS": {"tab": "5B21B6", "hdr": "4C1D95", "dat": "F5F3FF", "bdr": "C4B5FD"},
    "GEO": {"tab": "92400E", "hdr": "78350F", "dat": "FEF3C7", "bdr": "FCD34D"},
    "DIV": {"tab": "0D1F38", "hdr": "374151", "dat": "F3F4F6", "bdr": "D1D5DB"},
}

C_NEUTRAL_HDR = "6B7280"   # gris neutre — en-tête colonne info
C_NEUTRAL_DAT = "F9FAFB"   # blanc cassé — donnée colonne info

C_CYAN       = "00FFFF"
C_DARK_NAVY  = "0D1F38"
C_NAVY       = "1E3A5F"
C_AMBER      = "F59E0B"
C_AMBER_LITE = "FFF8E1"
C_MDC_BG     = "EBF4FF"
C_ET_BG      = "FFF7ED"
C_SECTION_BG = "1E3A5F"
C_KEY_BG     = "F1F5F9"
C_WHITE      = "FFFFFF"
C_GRAY       = "94A3B8"
C_BORDER     = "CBD5E1"

# ── Helpers globaux ───────────────────────────────────────────────────────────
def fill(h):        return PatternFill("solid", fgColor=h)
def bdr(color):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left",   vertical="center", indent=1)


def set_tab(ws, group_key):
    ws.sheet_properties.tabColor = GROUPS[group_key]["tab"]


def define_pk_named_range(wb):
    """
    Définit la plage nommée LISTE_PK au niveau du classeur.

    Formule : OFFSET depuis A2 de AXE_PROFIL_LONG, hauteur = nb de PK réels.
    Visible dans Excel → Formules → Gestionnaire de noms → LISTE_PK.
    Modifiable par l'utilisateur pour tout projet, quelle que soit la taille.
    """
    dn = DefinedName(
        name="LISTE_PK",
        attr_text="OFFSET(AXE_PROFIL_LONG!$A$2,0,0,COUNTA(AXE_PROFIL_LONG!$A:$A)-1,1)",
        comment="Liste dynamique de tous les PK du projet — ne pas supprimer",
    )
    wb.defined_names["LISTE_PK"] = dn


def add_pk_dv(ws, cell_range):
    """
    Dropdown PK référençant la plage nommée LISTE_PK (définie au niveau classeur).
    Aucune limite de taille — s'adapte à n'importe quel projet.
    """
    dv = DataValidation(
        type="list",
        formula1="LISTE_PK",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="PK invalide",
        error="Choisir un PK dans la liste (onglet AXE_PROFIL_LONG)"
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_list_dv(ws, cell_range, choices):
    """Dropdown liste fixe (ex: G,D ou AXLE,FIL_EAU)."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(choices)}"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def sheet_setup(ws, group_key, banner_text, headers, widths, col_types=None):
    """
    Prépare un onglet :
      - Tab color
      - Ligne 1 : bandeau
      - Ligne 2 : en-têtes colonnes (couleur groupe si "c", neutre si "i")
    Retourne le numéro de la première ligne de données (3).
    """
    ctypes = col_types or ["c"] * len(headers)
    set_tab(ws, group_key)
    g   = GROUPS[group_key]
    nc  = len(headers)

    # Bandeau
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws.cell(1, 1, banner_text)
    c.fill = fill(g["hdr"])
    c.font = Font(color="FFFFFF", bold=True, size=9)
    c.alignment = AL_L
    ws.row_dimensions[1].height = 18

    # En-têtes : couleur selon type colonne
    for ci, (h, w, ct) in enumerate(zip(headers, widths, ctypes), 1):
        hdr_clr = g["hdr"] if ct == "c" else C_NEUTRAL_HDR
        c = ws.cell(2, ci, h)
        c.fill = fill(hdr_clr)
        c.font = Font(color="FFFFFF", bold=True, size=8)
        c.alignment = AL_C
        c.border = bdr(g["bdr"] if ct == "c" else C_NEUTRAL_HDR)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 34
    return 3


def data_row(ws, ri, vals, group_key, bold_cols=None, alt=False, col_types=None):
    """Écrit une ligne de données avec le style du groupe.
    col_types : liste "c"/"i" par colonne — "c" = couleur groupe, "i" = neutre.
    """
    g         = GROUPS[group_key]
    bold_cols = bold_cols or []
    ctypes    = col_types or ["c"] * len(vals)
    for ci, (val, ct) in enumerate(zip(vals, ctypes), 1):
        if alt:
            bg_clr = "FFFFFF"
        else:
            bg_clr = g["dat"] if ct == "c" else C_NEUTRAL_DAT
        c = ws.cell(ri, ci, val)
        c.fill = fill(bg_clr)
        c.font = Font(size=9, bold=(ci in bold_cols), color="111827")
        c.alignment = AL_C
        c.border = bdr(g["bdr"] if ct == "c" else C_NEUTRAL_HDR)
    ws.row_dimensions[ri].height = 15


def legend_row(ws, ri, text, group_key):
    """Ligne de légende / note (fond très clair, italique)."""
    g = GROUPS[group_key]
    nc = ws.max_column or 1
    ws.merge_cells(f"A{ri}:{get_column_letter(nc)}{ri}")
    c = ws.cell(ri, 1, text)
    c.fill = fill("F8FAFC")
    c.font = Font(color="6B7280", size=7, italic=True)
    c.alignment = AL_L
    ws.row_dimensions[ri].height = 13


def make_logo_placeholder(role_label, company_name, bg_hex, border_hex):
    """Génère un image PNG placeholder logo (220×54 px) pour MDC ou ET."""
    def _h(h): return tuple(int(h[i:i+2], 16)/255 for i in (0, 2, 4))
    bg, brd = _h(bg_hex), _h(border_hex)
    fig, ax = plt.subplots(figsize=(2.8, 0.7), dpi=48)
    ax.set_facecolor(bg); fig.patch.set_facecolor(bg)
    ax.text(0.5, 0.65, f"[ LOGO  {role_label} ]",
            ha="center", va="center", fontsize=9, fontweight="bold",
            color=brd, transform=ax.transAxes)
    ax.text(0.5, 0.22, company_name,
            ha="center", va="center", fontsize=7, style="italic",
            color=brd, transform=ax.transAxes)
    for sp in ax.spines.values():
        sp.set_edgecolor(brd); sp.set_linewidth(2)
    ax.set_xticks([]); ax.set_yticks([])
    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.06)
    plt.close(fig)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
#  GEN_GENERALE
# ═══════════════════════════════════════════════════════════════════════════════
def build_gen_generale(ws):
    set_tab(ws, "GEN")

    thin = Side(style="thin", color=C_BORDER)
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LEFT_I = Alignment(horizontal="left",   vertical="center", indent=1, wrap_text=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 3
    row = 1

    def kfill(h):  return PatternFill("solid", fgColor=h)

    # Bandeau titre
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1, "RECEPTA  —  INFORMATIONS GÉNÉRALES DU PROJET")
    c.fill = kfill(C_DARK_NAVY)
    c.font = Font(color=C_CYAN, bold=True, size=14)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 36; row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1, "Onglet réservé — ne pas modifier les clés de la colonne A")
    c.fill = kfill("0F2744")
    c.font = Font(color=C_GRAY, size=8, italic=True)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 14; row += 2

    def section_hdr(label, bg=C_SECTION_BG):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, f"  {label}")
        c.fill = kfill(bg); c.font = Font(color=C_WHITE, bold=True, size=10)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 22; row += 1

    def kv(key, val, kbg=C_KEY_BG, vbg=C_WHITE, bold_v=False, h=18):
        nonlocal row
        ck = ws.cell(row, 1, key)
        ck.fill = kfill(kbg); ck.font = Font(color=C_NAVY, size=9, bold=True)
        ck.alignment = LEFT_I; ck.border = BORDER
        cv = ws.cell(row, 2, val)
        cv.fill = kfill(vbg); cv.font = Font(color="111827", size=10, bold=bold_v)
        cv.alignment = LEFT_I; cv.border = BORDER
        ws.row_dimensions[row].height = h; r = row; row += 1; return r

    def spacer(h=6):
        nonlocal row
        ws.row_dimensions[row].height = h; row += 1

    def logo_row(key_text, img_buf, kbg, vbg, h=56):
        nonlocal row
        ck = ws.cell(row, 1, key_text)
        ck.fill = kfill(kbg); ck.font = Font(color=C_NAVY, size=8, bold=True)
        ck.alignment = LEFT_I; ck.border = BORDER
        cv = ws.cell(row, 2, "")
        cv.fill = kfill(vbg); cv.border = BORDER
        ws.row_dimensions[row].height = h
        img = XLImage(img_buf)
        img.width = 220; img.height = h - 6
        ws.add_image(img, f"B{row}")
        row += 1

    # 1. Identification
    section_hdr("1. IDENTIFICATION DU PROJET")
    kv("GEN_NOM_PROJET",   "RN7 — Réhabilitation et élargissement 2×2 voies", bold_v=True, h=22)
    kv("GEN_INTITULE",     "Projet d'Aménagement et de Pavage Continu (PAPC) 2026")
    kv("GEN_LOCALISATION", "Région de Bobo-Dioulasso — Axe Bobo / Diébougou")
    kv("GEN_PK_DEBUT",     "0+000"); kv("GEN_PK_FIN", "3+000")
    kv("GEN_LONGUEUR_KM",  "3.000")
    kv("GEN_DESCRIPTION",
       "Réhabilitation de la chaussée existante avec élargissement à 2×2 voies, "
       "mise en place du réseau d'assainissement longitudinal et aménagement des carrefours.",
       h=48)
    spacer()

    # 2. MDC
    section_hdr("2. MISSION DE CONTRÔLE (MDC)", bg="1A4B8C")
    kv("GEN_MDC_NOM",     "OPTILAB Ingénierie", kbg="DBEAFE", vbg=C_MDC_BG, bold_v=True, h=22)
    kv("GEN_MDC_CONTACT", "Jean DUPONT — Ingénieur topographe", kbg="DBEAFE", vbg=C_MDC_BG)
    kv("GEN_MDC_EMAIL",   "j.dupont@optilab.fr", kbg="DBEAFE", vbg=C_MDC_BG)
    kv("GEN_MDC_TEL",     "+226 20 XX XX XX",    kbg="DBEAFE", vbg=C_MDC_BG)
    logo_row("GEN_MDC_LOGO",
             make_logo_placeholder("MDC", "OPTILAB Ingénierie", "EBF4FF", "0F4C81"),
             kbg="DBEAFE", vbg=C_MDC_BG)
    spacer()

    # 3. ET
    section_hdr("3. ENTREPRISE DE TRAVAUX (ET)", bg="92400E")
    kv("GEN_ET_NOM",     "BTP SAHEL SARL",      kbg="FED7AA", vbg=C_ET_BG, bold_v=True, h=22)
    kv("GEN_ET_CONTACT", "Moussa TRAORÉ — Chef de chantier", kbg="FED7AA", vbg=C_ET_BG)
    kv("GEN_ET_EMAIL",   "m.traore@btpsahel.bf", kbg="FED7AA", vbg=C_ET_BG)
    kv("GEN_ET_TEL",     "+226 25 XX XX XX",     kbg="FED7AA", vbg=C_ET_BG)
    logo_row("GEN_ET_LOGO",
             make_logo_placeholder("ET", "BTP SAHEL SARL", "FFF7ED", "92400E"),
             kbg="FED7AA", vbg=C_ET_BG)
    spacer()

    # 4. Autres acteurs
    section_hdr("4. AUTRES ACTEURS", bg="374151")
    kv("GEN_MAITRE_OUVRAGE",    "Ministère des Infrastructures et du Désenclavement")
    kv("GEN_BAILLEUR",          "Banque Mondiale — Projet Transport BF-P168426")
    kv("GEN_INGENIEUR_CONSEIL", "Cabinet SETEC International")
    kv("GEN_LABORATOIRE",       "LNBTP — Laboratoire National du BTP")
    spacer()

    # 5. Paramètres généraux
    section_hdr("5. PARAMÈTRES GÉNÉRAUX", bg="374151")
    kv("GEN_TYPE_CHAUSSEE",   "Chaussée bitumineuse flexible — 2×2 voies + TPC")
    kv("GEN_TOLERANCE_CM",    "2.0", bold_v=True)
    kv("GEN_MODE_DEFAUT",     "assainissement")
    kv("GEN_DEVERS_CHAUSSEE", "2.5 %")
    kv("GEN_VITESSE_PROJET",  "100 km/h")
    spacer()

    # Pied
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1, "RECEPTA by OPTILAB — Ne pas modifier les clés (colonne A)")
    c.fill = kfill(C_DARK_NAVY)
    c.font = Font(color=C_GRAY, size=7, italic=True)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 14

    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 90


# ═══════════════════════════════════════════════════════════════════════════════
#  AXE_PROFIL_LONG
# ═══════════════════════════════════════════════════════════════════════════════
def build_axe_profil_long(ws, df_g, meta):
    """AXE_PROFIL_LONG : PK | AXE_Z_axe — profil en long."""
    build_profil_long_sheet(ws, df_g, meta)   # réutilise gen_modeles_v3
    set_tab(ws, "AXE")
    # Renommer l'en-tête col B avec le préfixe AXE_
    ws.cell(1, 2, "AXE_Z_axe (m NGF)")


# ═══════════════════════════════════════════════════════════════════════════════
#  AXE_SECTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def build_axe_sections(ws, meta):
    """AXE_SECTIONS : épaisseurs signées par section (terrassement)."""
    HEADERS = ["Num", "PK_debut", "PK_fin", "Nom_Section",
               "AXE_e_BB", "AXE_e_GB4", "AXE_e_GNT", "AXE_e_FF", "AXE_e_PST"]
    WIDTHS  = [5, 11, 11, 24, 11, 11, 11, 11, 11]
    # i = identifiant/info  |  c = valeur utilisée dans les calculs
    CTYPES  = ["i", "c", "c", "i", "c", "c", "c", "c", "c"]
    sheet_setup(ws, "AXE",
                "AXE_SECTIONS — Épaisseurs signées par tronçon  "
                "| Négatif = sous la référence (Z_axe) | PK_debut ≤ PK < PK_fin",
                HEADERS, WIDTHS, col_types=CTYPES)

    # Dropdown PK sur les deux colonnes PK
    add_pk_dv(ws, "B3:B200")
    add_pk_dv(ws, "C3:C200")

    # Données exemples P1
    pk_fin_str = f"{meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"
    rows = [
        (1, "0+000", pk_fin_str, "Section principale",
         meta.get("ep_bb", -0.060), meta.get("ep_gb4"), meta.get("ep_gnt", -0.150),
         meta.get("ep_ff", -0.200), None),
    ]
    # Forcer les signes négatifs si positifs (les anciens meta stockaient unsigned)
    def signed(v):
        if v is None: return None
        return -abs(v) if v != 0 else 0.0

    for i, row in enumerate(rows):
        num, pkd, pkf, nom, bb, gb4, gnt, ff, pst = row
        vals = [num, pkd, pkf, nom,
                signed(bb), signed(gb4), signed(gnt), signed(ff), signed(pst)]
        data_row(ws, 3 + i, vals, "AXE", bold_cols=[1, 2], col_types=CTYPES)

    # 2 lignes gabarit
    for ri in [4, 5]:
        data_row(ws, ri, [""] * 9, "AXE", alt=True, col_types=CTYPES)

    legend_row(ws, 6,
        "Légende : e_BB négatif = 6 cm sous Z_axe  |  e_GB4 vide = couche absente  "
        "|  AXE_e_PST = couche de forme traitée (optionnelle)", "AXE")
    ws.freeze_panes = "D3"


# ═══════════════════════════════════════════════════════════════════════════════
#  TER_PROFIL_TYPE
# ═══════════════════════════════════════════════════════════════════════════════
def build_ter_profil_type(ws, meta):
    """TER_PROFIL_TYPE : points de mesure terrassement, de gauche vers la droite."""
    HEADERS = ["Num", "PK_debut", "PK_fin", "TER_Cote", "TER_Ordre",
               "TER_Label", "TER_Dist_Axe_m", "TER_Pente_pct"]
    WIDTHS  = [5, 11, 11, 8, 7, 22, 14, 13]
    CTYPES  = ["i", "c", "c", "c", "i", "i", "c", "c"]
    sheet_setup(ws, "TER",
                "TER_PROFIL_TYPE — Points de mesure terrassement par section  "
                "| TER_Cote : G=Gauche / D=Droit / AXE=axe  "
                "| Dist_Axe_m : distance absolue depuis l'axe (m)  "
                "| Pente_pct : dévers chaussée (%)",
                HEADERS, WIDTHS, col_types=CTYPES)

    add_pk_dv(ws, "B3:B200")
    add_pk_dv(ws, "C3:C200")
    add_list_dv(ws, "D3:D200", ["G", "D", "AXE"])

    pk_fin_str = f"{meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"
    pg = round(meta["pente_g"] * 100, 2)
    pd = round(meta["pente_d"] * 100, 2)
    dg = meta["dist_g"]
    dd = meta["dist_d"]

    # Points P1 : RN 2×2 voies — 4 points par côté
    pts = [
        # Gauche
        (1, "0+000", pk_fin_str, "G", 1, "TG_Bord_Voie1",    round(dg * 3.50/dg,  2), pg),
        (1, "0+000", pk_fin_str, "G", 2, "TG_Bord_Voie2",    round(dg * 7.00/dg,  2), pg),
        (1, "0+000", pk_fin_str, "G", 3, "TG_BAU",           round(dg * 10.0/dg,  2), pg),
        (1, "0+000", pk_fin_str, "G", 4, "TG_Bord_Chaussee", dg,                       pg),
        # Droit
        (1, "0+000", pk_fin_str, "D", 1, "TD_Bord_Voie1",    round(dd * 3.50/dd,  2), pd),
        (1, "0+000", pk_fin_str, "D", 2, "TD_Bord_Voie2",    round(dd * 7.00/dd,  2), pd),
        (1, "0+000", pk_fin_str, "D", 3, "TD_BAU",           round(dd * 10.0/dd,  2), pd),
        (1, "0+000", pk_fin_str, "D", 4, "TD_Bord_Chaussee", dd,                       pd),
    ]
    for i, row in enumerate(pts):
        data_row(ws, 3 + i, list(row), "TER",
                 bold_cols=[2, 6], alt=(i % 2 == 1), col_types=CTYPES)

    ri_after = 3 + len(pts)
    data_row(ws, ri_after,     [""] * 8, "TER", alt=True,  col_types=CTYPES)
    data_row(ws, ri_after + 1, [""] * 8, "TER", alt=False, col_types=CTYPES)
    legend_row(ws, ri_after + 2,
        "Ajouter des lignes pour chaque point de mesure.  "
        "Même Section = même Num.  Nouvelle section = nouveau bloc avec PK_debut différent.", "TER")
    ws.freeze_panes = "F3"


# ═══════════════════════════════════════════════════════════════════════════════
#  ASG_SECTIONS  /  ASD_SECTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def build_assain_sections(ws, group_key, meta, h_canal, e_dalle, dim_canal, prefix):
    """
    Construit ASG_SECTIONS ou ASD_SECTIONS.

    Structure :
      Cols A–H  : identité de la section + géométrie de référence
      Cols I+   : un offset signé par élément à réceptionner (depuis Z_fil_eau)
                  Ajouter librement de nouvelles colonnes {prefix}xxx à la suite.

    Convention des offsets (depuis Z_fil_eau) :
      positif (+) = élément AU-DESSUS du fil d'eau
      négatif (−) = élément EN DESSOUS du fil d'eau
    """
    nom_lot = "Assainissement Gauche" if "ASG" in group_key else "Assainissement Droit"

    # ── Colonnes fixes (géométrie de section) ────────────────────────────────
    HEADERS_GEO = [
        "Num", "PK_debut", "PK_fin", "Nom_Section",
        f"{prefix}Ref_Mode",       # AXLE = calculé depuis Z_axe | FIL_EAU = ASS_LONG
        f"{prefix}Dist_Axe_m",     # distance transversale depuis l'axe (m)
        f"{prefix}Pente_trans_pct",# dévers transversal (%)
        f"{prefix}Dim_canal",      # description canal (info, ex: 100×100)
    ]
    WIDTHS_GEO  = [5, 11, 11, 20, 11, 13, 17, 14]
    CTYPES_GEO  = ["i", "c", "c", "i", "c", "c", "c", "i"]

    # ── Colonnes éléments (offsets depuis Z_fil_eau) ─────────────────────────
    HEADERS_ELEM = [
        f"{prefix}BP",             # Béton de propreté       — en-dessous radier (−)
        f"{prefix}Radier",         # Haut du radier          — = fil d'eau (0)
        f"{prefix}Fil_eau",        # Fil d'eau (référence)   — 0.000
        f"{prefix}Tablier",        # Sous-face dalle de couv — au-dessus (+H+e)
        f"{prefix}Tete_can",       # Tête caniveau / niveau fini — idem Tablier
        f"{prefix}Abaissement",    # Seuil / abaissement     — optionnel
    ]
    WIDTHS_ELEM  = [11, 11, 11, 11, 11, 13]
    CTYPES_ELEM  = ["c", "c", "c", "c", "c", "c"]

    HEADERS = HEADERS_GEO + HEADERS_ELEM
    WIDTHS  = WIDTHS_GEO  + WIDTHS_ELEM
    CTYPES  = CTYPES_GEO  + CTYPES_ELEM

    sheet_setup(ws, group_key,
                f"{group_key}_SECTIONS — {nom_lot} par section  "
                f"| Ref_Mode : AXLE = Z_axe + pente transv.  |  FIL_EAU = ASS_LONG  "
                f"| Colonnes I+ : offsets depuis Z_fil_eau (+ au-dessus, − en dessous)",
                HEADERS, WIDTHS, col_types=CTYPES)

    add_pk_dv(ws, "B3:B200")
    add_pk_dv(ws, "C3:C200")
    add_list_dv(ws, "E3:E200", ["AXLE", "FIL_EAU"])

    pk_fin_str = f"{meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"
    pente = round((meta["pente_g"] if "ASG" in group_key else meta["pente_d"]) * 100, 2)
    dist  = meta["dist_g"] if "ASG" in group_key else meta["dist_d"]

    # Calcul des offsets pour l'exemple P1
    # Z_fil_eau est la référence (0). Toutes les cotes = Z_fil_eau + offset.
    e_radier    = 0.100   # épaisseur du radier (m)
    e_bp        = 0.200   # épaisseur béton de propreté (m)
    off_bp      = round(-(e_radier + e_bp),   3)   # -0.300
    off_radier  = 0.000                             #  0.000  (haut radier = fil d'eau)
    off_fileau  = 0.000                             #  0.000  (référence)
    off_tablier = round(+(h_canal + e_dalle),  3)   # ex: +1.150 pour H=1.0, e=0.15
    off_tete    = off_tablier                       # tête caniveau = dessous Tablier
    off_abais   = None                              # à renseigner si besoin

    row1 = [
        1, "0+000", pk_fin_str, f"Canal {dim_canal}",
        "AXLE", dist, pente, dim_canal,
        off_bp, off_radier, off_fileau, off_tablier, off_tete, off_abais,
    ]
    data_row(ws, 3, row1, group_key, bold_cols=[1, 2, 5, 9], col_types=CTYPES)

    for ri in [4, 5]:
        data_row(ws, ri, [""] * len(HEADERS), group_key, alt=(ri % 2 == 0), col_types=CTYPES)

    legend_row(ws, 6,
        f"Mode AXLE : Z_fil_eau = Z_axe − Dist×Pente/100  |  "
        f"Z_element = Z_fil_eau + offset_colonne  |  "
        f"Ajouter {prefix}xxx en colonne {chr(73 + len(HEADERS_ELEM))}+ pour tout élément supplémentaire",
        group_key)
    ws.freeze_panes = "E3"


# ═══════════════════════════════════════════════════════════════════════════════
#  ASS_LONG
# ═══════════════════════════════════════════════════════════════════════════════
def build_ass_long(ws, df_g, meta):
    """
    ASS_LONG : cotes fil d'eau par PK, par côté.
    Utilisé uniquement si Ref_Mode = FIL_EAU dans ASG/ASD_SECTIONS.
    En mode AXLE, cet onglet est ignoré par l'app.
    """
    HEADERS = ["PK", "AG_Z_fil_eau", "AD_Z_fil_eau"]
    WIDTHS  = [10, 16, 16]
    sheet_setup(ws, "ASS",
                "ASS_LONG — Profil en long fil d'eau  "
                "| Actif uniquement si Ref_Mode = FIL_EAU dans ASG_ ou ASD_SECTIONS",
                HEADERS, WIDTHS)

    # Calculer les cotes fil d'eau depuis les paramètres P1
    Z0, PL   = meta["z0"], float(meta["pente_long"].replace("%", "")) / 100
    dg, dd   = meta["dist_g"], meta["dist_d"]
    pg, pd_  = meta["pente_g"], meta["pente_d"]
    # P1 : H_canal_G=1.00 e_dalle_G=0.15 / H_canal_D=0.80 e_dalle_D=0.12
    h_g, e_g = meta["h_canal_g"], 0.15
    h_d, e_d = meta["h_canal_d"], 0.12

    pks = df_g["PK"].tolist()
    for i, pk_str in enumerate(pks):
        pk_m = int(pk_str.split("+")[0]) * 1000 + int(pk_str.split("+")[1])
        z    = Z0 + pk_m * PL
        z_brd_g = round(z - dg * pg, 3)
        z_brd_d = round(z - dd * pd_, 3)
        z_fe_g  = round(z_brd_g - h_g - e_g, 3)
        z_fe_d  = round(z_brd_d - h_d - e_d, 3)
        data_row(ws, 3 + i, [pk_str, z_fe_g, z_fe_d], "ASS",
                 bold_cols=[1], alt=(i % 2 == 0))

    legend_row(ws, 3 + len(pks),
        "Ces valeurs ont été pré-calculées à titre d'exemple (mode AXLE).  "
        "En production : renseigner les vraies cotes fil d'eau issues du plan hydraulique.", "ASS")
    ws.freeze_panes = "B3"


# ═══════════════════════════════════════════════════════════════════════════════
#  GEO_COORDONNEES
# ═══════════════════════════════════════════════════════════════════════════════
def build_geo_coordonnees(ws, df_g, meta):
    """GEO_COORDONNEES : coordonnées X Y Z + cap pour chaque PK de l'axe."""
    HEADERS = ["PK", "XY_X", "XY_Y", "XY_Z", "XY_Cap_deg"]
    WIDTHS  = [10, 16, 16, 14, 12]
    sheet_setup(ws, "GEO",
                "GEO_COORDONNEES — Coordonnées absolues de l'axe principal  "
                "| XY_Cap_deg : azimut en degrés décimaux (0° = Nord, 90° = Est)",
                HEADERS, WIDTHS)

    Z0, PL = meta["z0"], float(meta["pente_long"].replace("%", "")) / 100
    cap = 47.3
    cap_rad = math.radians(cap)
    x0, y0 = 612340.250, 1432180.600

    pks = df_g["PK"].tolist()
    for i, pk_str in enumerate(pks):
        pk_m = int(pk_str.split("+")[0]) * 1000 + int(pk_str.split("+")[1])
        x    = round(x0 + pk_m * math.sin(cap_rad), 3)
        y    = round(y0 + pk_m * math.cos(cap_rad), 3)
        z    = round(Z0 + pk_m * PL, 3)
        data_row(ws, 3 + i, [pk_str, x, y, z, cap], "GEO",
                 bold_cols=[1], alt=(i % 2 == 0))

    legend_row(ws, 3 + len(pks),
        "Coordonnées exemple (azimut 47.3°, X0=612340.250, Y0=1432180.600).  "
        "Remplacer par les vraies coordonnées du projet.", "GEO")
    ws.freeze_panes = "B3"


# ═══════════════════════════════════════════════════════════════════════════════
#  GEO_AXES_PARALLELES
# ═══════════════════════════════════════════════════════════════════════════════
def build_geo_axes_paralleles(ws, meta):
    """GEO_AXES_PARALLELES : définition des axes parallèles à l'axe principal."""
    HEADERS = ["Num", "PK_debut", "PK_fin", "XY_Label",
               "XY_Cote", "XY_Dist_m", "XY_Description"]
    WIDTHS  = [5, 11, 11, 24, 8, 11, 32]
    CTYPES  = ["i", "c", "c", "c", "c", "c", "i"]
    sheet_setup(ws, "GEO",
                "GEO_AXES_PARALLELES — Axes parallèles à l'axe principal par section  "
                "| XY_Dist_m : distance perpendiculaire depuis l'axe (m)  "
                "| XY_Cote : G=Gauche / D=Droit",
                HEADERS, WIDTHS, col_types=CTYPES)

    add_pk_dv(ws, "B3:B200")
    add_pk_dv(ws, "C3:C200")
    add_list_dv(ws, "E3:E200", ["G", "D"])

    pk_fin_str = f"{meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"
    dg, dd = meta["dist_g"], meta["dist_d"]

    axes = [
        (1, "0+000", pk_fin_str, "TG_Bord_Chaussee", "G", dg,        "Bord chaussée gauche"),
        (2, "0+000", pk_fin_str, "AG_Canal_Axe",      "G", dg + 0.20, "Axe caniveau gauche"),
        (3, "0+000", pk_fin_str, "AG_Bord_Propriete", "G", dg + 3.00, "Limite emprise gauche"),
        (4, "0+000", pk_fin_str, "TD_Bord_Chaussee",  "D", dd,        "Bord chaussée droit"),
        (5, "0+000", pk_fin_str, "AD_Canal_Axe",      "D", dd + 0.20, "Axe caniveau droit"),
        (6, "0+000", pk_fin_str, "AD_Bord_Propriete", "D", dd + 3.00, "Limite emprise droite"),
    ]
    for i, row in enumerate(axes):
        data_row(ws, 3 + i, list(row), "GEO",
                 bold_cols=[1, 4], alt=(i % 2 == 1), col_types=CTYPES)

    ri_after = 3 + len(axes)
    data_row(ws, ri_after, [""] * 7, "GEO", alt=True, col_types=CTYPES)
    legend_row(ws, ri_after + 1,
        "Ajouter une ligne par axe parallèle.  "
        "Si la distance change à un PK donné, ajouter une nouvelle ligne avec le même label "
        "et un PK_debut différent.", "GEO")
    ws.freeze_panes = "D3"


# ═══════════════════════════════════════════════════════════════════════════════
#  DIV_IMPREVUS
# ═══════════════════════════════════════════════════════════════════════════════
def build_div_imprevus(ws, df_g, df_d, meta):
    build_imprevus_sheet(ws, df_g, df_d, meta)
    set_tab(ws, "DIV")


# ═══════════════════════════════════════════════════════════════════════════════
#  DIV_LEGENDE
# ═══════════════════════════════════════════════════════════════════════════════
def build_div_legende(ws, df_g, df_d, meta):
    set_tab(ws, "DIV")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 72

    img_buf = draw_profil(meta)
    img = XLImage(img_buf)
    img.width = 980; img.height = 360
    ws.add_image(img, "B2")
    for r in range(2, 25):
        ws.row_dimensions[r].height = 17

    def _c(row, txt, bg=None, fg="FFFFFF", bold=False, sz=10):
        c = ws.cell(row, 2, txt)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, bold=bold, size=sz)
        else:
            c.font = Font(size=sz or 9)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = sz + 5 if sz else 8

    r = 25
    _c(r, "SPECIFICATIONS DU PROJET", "0D1F38", "00FFFF", True, 12); r += 1
    for k, v in [
        ("Titre",        meta["titre"]),
        ("Devers",       meta["devers"]),
        ("Largeurs",     meta["largeur"]),
        ("Couches",      meta["couches"]),
        ("Pente long.",  meta["pente_long"]),
        ("Cote axe Z0",  f"{meta['z0']:.3f} m NGF"),
        ("PK debut/fin", f"0+000 / {meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"),
        ("Pas PK",       f"{meta['pas']} m"),
    ]:
        _c(r, f"  {k+':':15s} {v}", "1E3A5F", "FFFFFF", False, 10); r += 1

    _c(r, "", sz=6); r += 1
    _c(r, "CONVENTION DES SIGNES — EPAISSEURS", "374151", "FFFFFF", True, 10); r += 1
    _c(r, "  Couche EN DESSOUS de la référence  → valeur NÉGATIVE  ex: AXE_e_BB = -0.060", "1E3A5F", "FFFFFF", False, 9); r += 1
    _c(r, "  Couche EN DESSUS  de la référence  → valeur POSITIVE  ex: AG_H_canal = +1.000", "1E3A5F", "FFFFFF", False, 9); r += 1
    _c(r, "  Z_element = Z_reference + somme des épaisseurs signées", "0F4C81", "00FFFF", True, 9); r += 1

    _c(r, "", sz=6); r += 1
    _c(r, "CONVENTION DES SECTIONS", "374151", "FFFFFF", True, 10); r += 1
    _c(r, "  PK_debut uniquement — section active = dernière dont PK_debut ≤ PK courant", "1E3A5F", "FFFFFF", False, 9); r += 1
    _c(r, "  Dropdowns sur PK_debut et PK_fin → liste depuis AXE_PROFIL_LONG", "1E3A5F", "FFFFFF", False, 9); r += 1

    _c(r, "", sz=6); r += 1
    _c(r, "CANAL GAUCHE (P1) — " + meta["cv_g"], "6B21A8", "FFFFFF", True, 10); r += 1
    _c(r, "  " + meta["assain_g"], "6B21A8", "FFFFFF", False, 9); r += 1
    _c(r, "CANAL DROIT  (P1) — " + meta["cv_d"], "9D174D", "FFFFFF", True, 10); r += 1
    _c(r, "  " + meta["assain_d"], "9D174D", "FFFFFF", False, 9); r += 1


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    df_g, df_d, meta = projet_1()
    wb = Workbook()

    # Plage nommée LISTE_PK — référencée par tous les dropdowns PK du classeur
    define_pk_named_range(wb)

    # 1. GEN_GENERALE
    ws = wb.active; ws.title = "GEN_GENERALE"
    build_gen_generale(ws)

    # 2. AXE_PROFIL_LONG
    ws = wb.create_sheet("AXE_PROFIL_LONG")
    build_axe_profil_long(ws, df_g, meta)

    # 3. AXE_SECTIONS
    ws = wb.create_sheet("AXE_SECTIONS")
    build_axe_sections(ws, meta)

    # 4. TER_PROFIL_TYPE
    ws = wb.create_sheet("TER_PROFIL_TYPE")
    build_ter_profil_type(ws, meta)

    # 5. ASG_SECTIONS
    ws = wb.create_sheet("ASG_SECTIONS")
    build_assain_sections(ws, "ASG", meta,
                          h_canal=meta["h_canal_g"], e_dalle=0.15,
                          dim_canal=meta["cv_g"], prefix="AG_")

    # 6. ASD_SECTIONS
    ws = wb.create_sheet("ASD_SECTIONS")
    build_assain_sections(ws, "ASD", meta,
                          h_canal=meta["h_canal_d"], e_dalle=0.12,
                          dim_canal=meta["cv_d"], prefix="AD_")

    # 7. ASS_LONG
    ws = wb.create_sheet("ASS_LONG")
    build_ass_long(ws, df_g, meta)

    # 8. GEO_COORDONNEES
    ws = wb.create_sheet("GEO_COORDONNEES")
    build_geo_coordonnees(ws, df_g, meta)

    # 9. GEO_AXES_PARALLELES
    ws = wb.create_sheet("GEO_AXES_PARALLELES")
    build_geo_axes_paralleles(ws, meta)

    # 10. DIV_IMPREVUS
    ws = wb.create_sheet("DIV_IMPREVUS")
    build_div_imprevus(ws, df_g, df_d, meta)

    # 11. DIV_LEGENDE
    ws = wb.create_sheet("DIV_LEGENDE")
    build_div_legende(ws, df_g, df_d, meta)

    wb.save(OUT_PATH)

    # Projet P1 pré-rempli : recharger depuis le fichier (buffers images fermés après save)
    from openpyxl import load_workbook as _lw
    wb2 = _lw(OUT_PATH)
    wb2["GEN_GENERALE"].cell(1, 1,
        "RECEPTA  —  P1 RN 2x2 VOIES + TPC  |  3 km  |  PROJET EXEMPLE")
    wb2.save(OUT_PROJET)

    names = [s.title for s in wb.worksheets]
    print(f"OK  {OUT_PATH}  (modele type vide)")
    print(f"OK  {OUT_PROJET}  (projet P1 pre-rempli)")
    print(f"    {len(names)} onglets : {' | '.join(names)}")
    print()
    print("Groupes et couleurs tabs :")
    groups = {
        "GEN_": "#374151 gris",
        "AXE_": "#1E3A5F navy",
        "TER_": "#166534 vert",
        "ASG_": "#6B21A8 violet",
        "ASD_": "#9D174D bordeaux",
        "ASS_": "#5B21B6 violet moyen",
        "GEO_": "#92400E marron",
        "DIV_": "#0D1F38 dark navy",
    }
    for pfx, col in groups.items():
        matching = [n for n in names if n.startswith(pfx)]
        if matching:
            print(f"  {pfx}  {col}  ->  {', '.join(matching)}")


if __name__ == "__main__":
    main()
