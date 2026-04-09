# -*- coding: utf-8 -*-
"""
RECEPTA — Génération des 4 fichiers modèles Excel
Dossier sortie : data/modeles_recepta/

Stratigraphie assainissement (de bas en haut) :
  Fond_Fouille
  + 0.20m → top Beton_Proprete   (couche de 20cm)
  + 0.10m → Radier_Interieur     (radier canal e=10cm)
  + H_canal (1.00 ou 0.60 ou n)  → dessous Tablier/Dalle
  + e_dalle                      → Tablier_Fini (affleurant bord chaussee)
"""
import io, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

OUT_DIR = "data/modeles_recepta"
os.makedirs(OUT_DIR, exist_ok=True)

# ── Styles ────────────────────────────────────────────────────
thin   = Side(style="thin",   color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def fill(h): return PatternFill("solid", fgColor=h)

PALETTE = {
    "pk":     ("1E3A5F", "EBF4FF"),
    "axe":    ("0F4C81", "E8F0F8"),
    "trav_g": ("1A6B3C", "EAF5EE"),
    "trav_d": ("155E2F", "E4F5E9"),
    "ass_g":  ("6B2D8B", "F3E8FA"),
    "ass_d":  ("8B2D6B", "FAE8F3"),
}

# ── Génération des PK ─────────────────────────────────────────
def gen_pks(pk_fin, pas):
    pks, pk_m = [], []
    pk = 0
    while pk <= pk_fin:
        pks.append(f"{pk // 1000}+{pk % 1000:03d}")
        pk_m.append(pk)
        pk += pas
    return pks, pk_m

# ── Calcul cotes assainissement (commun) ──────────────────────
def assain_cotes(z_brd, h_canal, e_dalle, e_radier=0.10, e_bp=0.20):
    """
    Retourne dict des cotes assainissement pour un côté.
    z_brd  = cote bord chaussée = Tablier_Fini
    h_canal = hauteur intérieure utile du caniveau
    e_dalle = épaisseur de la dalle de couverture (tablier)
    e_radier= épaisseur du radier (fond canal) = 0.10m par défaut
    e_bp    = épaisseur béton de propreté = 0.20m par défaut
    """
    tablier    = round(z_brd, 3)
    rad_int    = round(tablier - e_dalle - h_canal, 3)   # intérieur fond canal
    bp_top     = round(rad_int - e_radier, 3)            # dessus béton propreté
    fond_fouille = round(bp_top - e_bp, 3)               # fond de fouille
    bp_haut    = round(z_brd + 0.20, 3)                  # bord propriété
    return {
        "Tablier_Fini":     tablier,
        "Radier_Int":       rad_int,
        "Beton_Proprete":   bp_top,
        "Fond_Fouille":     fond_fouille,
        "Bord_Propriete":   bp_haut,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 1 — Route Nationale 2×2 voies + TPC  (3 km)
#  Caniveau G : 100×100  |  Caniveau D : 80×80
# ═══════════════════════════════════════════════════════════════
def projet_1():
    Z0=102.500; PL=0.012; PK_FIN=3000; PAS=25
    PENTE_G=PENTE_D=0.025
    EP_BB=0.06; EP_GB4=0.10; EP_GNT=0.15; EP_FOND=0.20
    DIST_G=DIST_D=8.0  # 2 voies 3.5m + BAU 1m

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        # Points chaussée G
        z_v1g  = round(z - 3.50 * PENTE_G, 3)
        z_v2g  = round(z - 7.00 * PENTE_G, 3)
        z_brdG = round(z - DIST_G * PENTE_G, 3)
        g_bb=round(z_brdG,3); g_gb4=round(z_brdG-EP_BB,3)
        g_gnt=round(z_brdG-EP_BB-EP_GB4,3); g_fond=round(z_brdG-EP_BB-EP_GB4-EP_GNT,3)
        ag = assain_cotes(z_brdG, h_canal=1.00, e_dalle=0.15)
        rows_g.append({"PK":lb,
            "AXE_BB":z, "AXE_GB4":round(z-EP_BB,3),
            "AXE_GNT":round(z-EP_BB-EP_GB4,3), "AXE_Fond_Forme":round(z-EP_BB-EP_GB4-EP_GNT,3),
            "G_Bord_Voie1":z_v1g, "G_Bord_Voie2":z_v2g,
            "G_BB_Roulement":g_bb, "G_GB4_Base":g_gb4,
            "G_GNT_SousBase":g_gnt, "G_Fond_Forme":g_fond,
            "G_Bordure_Finie":round(z_brdG-0.02,3),
            "G_Tablier_Fini_100x100":  ag["Tablier_Fini"],
            "G_Radier_Interieur":      ag["Radier_Int"],
            "G_Beton_Proprete":        ag["Beton_Proprete"],
            "G_Fond_Fouille":          ag["Fond_Fouille"],
            "G_Bord_Propriete":        ag["Bord_Propriete"],
        })
        # Points chaussée D
        z_v1d  = round(z - 3.50 * PENTE_D, 3)
        z_v2d  = round(z - 7.00 * PENTE_D, 3)
        z_brdD = round(z - DIST_D * PENTE_D, 3)
        d_bb=round(z_brdD,3); d_gb4=round(z_brdD-EP_BB,3)
        d_gnt=round(z_brdD-EP_BB-EP_GB4,3); d_fond=round(z_brdD-EP_BB-EP_GB4-EP_GNT,3)
        ad = assain_cotes(z_brdD, h_canal=0.80, e_dalle=0.12)
        rows_d.append({"PK":lb,
            "AXE_BB":z, "AXE_GB4":round(z-EP_BB,3),
            "AXE_GNT":round(z-EP_BB-EP_GB4,3), "AXE_Fond_Forme":round(z-EP_BB-EP_GB4-EP_GNT,3),
            "D_Bord_Voie1":z_v1d, "D_Bord_Voie2":z_v2d,
            "D_BB_Roulement":d_bb, "D_GB4_Base":d_gb4,
            "D_GNT_SousBase":d_gnt, "D_Fond_Forme":d_fond,
            "D_Bordure_Finie":round(z_brdD-0.02,3),
            "D_Tablier_Fini_80x80":    ad["Tablier_Fini"],
            "D_Radier_Interieur":      ad["Radier_Int"],
            "D_Beton_Proprete":        ad["Beton_Proprete"],
            "D_Fond_Fouille":          ad["Fond_Fouille"],
            "D_Bord_Propriete":        ad["Bord_Propriete"],
        })
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre":"RN 2x2 voies + TPC", "devers":"G:2.5% | D:2.5%",
        "largeur":"2x(2x3.50m + BAU 1.0m) | dist axe/bord = 8.0m",
        "couches":"BB 6cm / GB4 10cm / GNT 15cm / FF 20cm",
        "assain_g":"Caniveau ferme 100x100 — H int=1.00m / dalle=15cm / radier=10cm / BP=20cm",
        "assain_d":"Caniveau ferme 80x80  — H int=0.80m / dalle=12cm / radier=10cm / BP=20cm",
        "pente_long":f"{PL*100:.1f}%", "z0":Z0, "pk_fin":PK_FIN, "pas":PAS,
        "cv_g":"100x100 (H=1.00m)", "cv_d":"80x80 (H=0.80m)",
        "pente_g":0.025, "pente_d":0.025, "dist_g":8.0, "dist_d":8.0,
        "h_canal_g":1.00, "h_canal_d":0.80,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 2 — Route Urbaine trottoirs + stationnement  (1.5 km)
#  Caniveau G : 60×60  |  Caniveau D : 40×40 (avaloir urbain)
# ═══════════════════════════════════════════════════════════════
def projet_2():
    Z0=98.200; PL=0.008; PK_FIN=1500; PAS=25
    PENTE_G=0.020; PENTE_D=0.030
    EP_BB=0.06; EP_TRAV=0.20; EP_FOND=0.25
    DIST_G=3.50; DIST_D=3.50

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        z_brdG = round(z - DIST_G * PENTE_G, 3)
        z_statG = round(z - 5.50 * PENTE_G, 3)
        z_trotG = round(z - 7.00 * PENTE_G + 0.10, 3)
        g_bb=round(z_brdG,3); g_tv=round(z_brdG-EP_BB,3); g_fd=round(z_brdG-EP_BB-EP_TRAV,3)
        ag = assain_cotes(z_brdG, h_canal=0.60, e_dalle=0.10)
        rows_g.append({"PK":lb,
            "AXE_BB":z, "AXE_Travaux":round(z-EP_BB,3), "AXE_Fond_Forme":round(z-EP_BB-EP_TRAV,3),
            "G_BB_Roulement":g_bb, "G_Travaux_Base":g_tv, "G_Fond_Forme":g_fd,
            "G_Bordure_Finie":round(z_brdG-0.02,3),
            "G_Bord_Stationnement":z_statG,
            "G_Trottoir_Fini":z_trotG,
            "G_Tablier_Fini_60x60":   ag["Tablier_Fini"],
            "G_Radier_Interieur":     ag["Radier_Int"],
            "G_Beton_Proprete":       ag["Beton_Proprete"],
            "G_Fond_Fouille":         ag["Fond_Fouille"],
            "G_Bord_Propriete":       ag["Bord_Propriete"],
        })
        z_brdD = round(z - DIST_D * PENTE_D, 3)
        z_trotD = round(z - 5.50 * PENTE_D + 0.12, 3)
        d_bb=round(z_brdD,3); d_tv=round(z_brdD-EP_BB,3); d_fd=round(z_brdD-EP_BB-EP_TRAV,3)
        ad = assain_cotes(z_brdD, h_canal=0.40, e_dalle=0.10)
        rows_d.append({"PK":lb,
            "AXE_BB":z, "AXE_Travaux":round(z-EP_BB,3), "AXE_Fond_Forme":round(z-EP_BB-EP_TRAV,3),
            "D_BB_Roulement":d_bb, "D_Travaux_Base":d_tv, "D_Fond_Forme":d_fd,
            "D_Bordure_Finie":round(z_brdD-0.02,3),
            "D_Trottoir_Fini":z_trotD,
            "D_Tablier_Fini_40x40":   ad["Tablier_Fini"],
            "D_Radier_Interieur":     ad["Radier_Int"],
            "D_Beton_Proprete":       ad["Beton_Proprete"],
            "D_Fond_Fouille":         ad["Fond_Fouille"],
            "D_Bord_Propriete":       ad["Bord_Propriete"],
        })
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre":"Route Urbaine trottoirs + stationnement", "devers":"G:2.0% | D:3.0%",
        "largeur":"G: chaussee 3.5m + stat 2.0m + trottoir 1.5m | D: chaussee 3.5m + trottoir 2.0m",
        "couches":"BB 6cm / Travaux 20cm / Fond de Forme 25cm",
        "assain_g":"Caniveau ferme 60x60 — H int=0.60m / dalle=10cm / radier=10cm / BP=20cm",
        "assain_d":"Caniveau ferme 40x40 — H int=0.40m / dalle=10cm / radier=10cm / BP=20cm",
        "pente_long":f"{PL*100:.1f}%", "z0":Z0, "pk_fin":PK_FIN, "pas":PAS,
        "cv_g":"60x60 (H=0.60m)", "cv_d":"40x40 (H=0.40m)",
        "pente_g":0.020, "pente_d":0.030, "dist_g":3.5, "dist_d":3.5,
        "h_canal_g":0.60, "h_canal_d":0.40,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 3 — Autoroute 2×3 voies + BAU  (5 km)
#  Caniveau G : 120×120  |  Caniveau D : 120×120
# ═══════════════════════════════════════════════════════════════
def projet_3():
    Z0=115.000; PL=0.006; PK_FIN=5000; PAS=25
    PENTE_G=PENTE_D=0.025
    EP_BB=0.06; EP_GB4=0.12; EP_GNT=0.20; EP_FOND=0.30; EP_PST=0.15
    DIST_G=DIST_D=15.25  # 3x3.75m + BAU3m + acc1m

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        z_v3g = round(z - 11.25 * PENTE_G, 3)
        z_bauG = round(z - 14.25 * PENTE_G, 3)
        z_accG = round(z - 15.25 * PENTE_G, 3)
        z_brdG = round(z_accG - 0.02, 3)
        g_bb=round(z_bauG,3); g_gb4=round(z_bauG-EP_BB,3)
        g_gnt=round(z_bauG-EP_BB-EP_GB4,3); g_fond=round(z_bauG-EP_BB-EP_GB4-EP_GNT,3)
        g_pst=round(g_fond-EP_PST,3)
        ag = assain_cotes(z_brdG, h_canal=1.20, e_dalle=0.18)
        rows_g.append({"PK":lb,
            "AXE_BB":z, "AXE_GB4":round(z-EP_BB,3), "AXE_GNT":round(z-EP_BB-EP_GB4,3),
            "AXE_Fond_Forme":round(z-EP_BB-EP_GB4-EP_GNT,3),
            "AXE_PST":round(z-EP_BB-EP_GB4-EP_GNT-EP_PST,3),
            "G_Bord_Voie3":z_v3g, "G_Bord_BAU":z_bauG, "G_Accotement":z_accG,
            "G_BB_Roulement":g_bb, "G_GB4_Base":g_gb4, "G_GNT_SousBase":g_gnt,
            "G_Fond_Forme":g_fond, "G_PST":g_pst,
            "G_Bordure_Finie":z_brdG,
            "G_Tablier_Fini_120x120":  ag["Tablier_Fini"],
            "G_Radier_Interieur":      ag["Radier_Int"],
            "G_Beton_Proprete":        ag["Beton_Proprete"],
            "G_Fond_Fouille":          ag["Fond_Fouille"],
            "G_Bord_Propriete":        ag["Bord_Propriete"],
        })
        z_v3d = round(z - 11.25 * PENTE_D, 3)
        z_bauD = round(z - 14.25 * PENTE_D, 3)
        z_accD = round(z - 15.25 * PENTE_D, 3)
        z_brdD = round(z_accD - 0.02, 3)
        d_bb=round(z_bauD,3); d_gb4=round(z_bauD-EP_BB,3)
        d_gnt=round(z_bauD-EP_BB-EP_GB4,3); d_fond=round(z_bauD-EP_BB-EP_GB4-EP_GNT,3)
        d_pst=round(d_fond-EP_PST,3)
        ad = assain_cotes(z_brdD, h_canal=1.20, e_dalle=0.18)
        rows_d.append({"PK":lb,
            "AXE_BB":z, "AXE_GB4":round(z-EP_BB,3), "AXE_GNT":round(z-EP_BB-EP_GB4,3),
            "AXE_Fond_Forme":round(z-EP_BB-EP_GB4-EP_GNT,3),
            "AXE_PST":round(z-EP_BB-EP_GB4-EP_GNT-EP_PST,3),
            "D_Bord_Voie3":z_v3d, "D_Bord_BAU":z_bauD, "D_Accotement":z_accD,
            "D_BB_Roulement":d_bb, "D_GB4_Base":d_gb4, "D_GNT_SousBase":d_gnt,
            "D_Fond_Forme":d_fond, "D_PST":d_pst,
            "D_Bordure_Finie":z_brdD,
            "D_Tablier_Fini_120x120":  ad["Tablier_Fini"],
            "D_Radier_Interieur":      ad["Radier_Int"],
            "D_Beton_Proprete":        ad["Beton_Proprete"],
            "D_Fond_Fouille":          ad["Fond_Fouille"],
            "D_Bord_Propriete":        ad["Bord_Propriete"],
        })
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre":"Autoroute 2x3 voies + BAU + accotement", "devers":"G:2.5% | D:2.5%",
        "largeur":"3x3.75m voies + BAU 3.0m + accotement 1.0m | dist axe/bord = 15.25m",
        "couches":"BB 6cm / GB4 12cm / GNT 20cm / FF 30cm / PST 15cm",
        "assain_g":"Caniveau ferme 120x120 — H int=1.20m / dalle=18cm / radier=10cm / BP=20cm",
        "assain_d":"Caniveau ferme 120x120 — H int=1.20m / dalle=18cm / radier=10cm / BP=20cm",
        "pente_long":f"{PL*100:.1f}%", "z0":Z0, "pk_fin":PK_FIN, "pas":PAS,
        "cv_g":"120x120 (H=1.20m)", "cv_d":"120x120 (H=1.20m)",
        "pente_g":0.025, "pente_d":0.025, "dist_g":15.25, "dist_d":15.25,
        "h_canal_g":1.20, "h_canal_d":1.20,
    }

# ═══════════════════════════════════════════════════════════════
#  PROJET 4 — Route Rurale secondaire 1×2 voies  (1 km)
#  Caniveau G : 50×50  |  Caniveau D : 50×50  (fossé maçonné)
# ═══════════════════════════════════════════════════════════════
def projet_4():
    Z0=85.750; PL=0.020; PK_FIN=1000; PAS=25
    PENTE_G=0.030; PENTE_D=0.040
    EP_BB=0.05; EP_GNT=0.20; EP_FOND=0.25
    DIST_G=4.0; DIST_D=4.0

    pks, pk_m = gen_pks(PK_FIN, PAS)
    rows_g, rows_d = [], []
    for lb, d in zip(pks, pk_m):
        z = round(Z0 + d * PL, 3)
        z_brdG = round(z - DIST_G * PENTE_G, 3)
        g_bb=round(z_brdG,3); g_gnt=round(z_brdG-EP_BB,3); g_fond=round(z_brdG-EP_BB-EP_GNT,3)
        ag = assain_cotes(z_brdG, h_canal=0.50, e_dalle=0.10)
        rows_g.append({"PK":lb,
            "AXE_BB":z, "AXE_GNT":round(z-EP_BB,3), "AXE_Fond_Forme":round(z-EP_BB-EP_GNT,3),
            "G_BB_Roulement":g_bb, "G_GNT_Base":g_gnt, "G_Fond_Forme":g_fond,
            "G_Bordure_Finie":round(z_brdG-0.02,3),
            "G_Tablier_Fini_50x50":   ag["Tablier_Fini"],
            "G_Radier_Interieur":     ag["Radier_Int"],
            "G_Beton_Proprete":       ag["Beton_Proprete"],
            "G_Fond_Fouille":         ag["Fond_Fouille"],
            "G_Bord_Propriete":       ag["Bord_Propriete"],
        })
        z_brdD = round(z - DIST_D * PENTE_D, 3)
        d_bb=round(z_brdD,3); d_gnt=round(z_brdD-EP_BB,3); d_fond=round(z_brdD-EP_BB-EP_GNT,3)
        ad = assain_cotes(z_brdD, h_canal=0.50, e_dalle=0.10)
        rows_d.append({"PK":lb,
            "AXE_BB":z, "AXE_GNT":round(z-EP_BB,3), "AXE_Fond_Forme":round(z-EP_BB-EP_GNT,3),
            "D_BB_Roulement":d_bb, "D_GNT_Base":d_gnt, "D_Fond_Forme":d_fond,
            "D_Bordure_Finie":round(z_brdD-0.02,3),
            "D_Tablier_Fini_50x50":   ad["Tablier_Fini"],
            "D_Radier_Interieur":     ad["Radier_Int"],
            "D_Beton_Proprete":       ad["Beton_Proprete"],
            "D_Fond_Fouille":         ad["Fond_Fouille"],
            "D_Bord_Propriete":       ad["Bord_Propriete"],
        })
    return pd.DataFrame(rows_g), pd.DataFrame(rows_d), {
        "titre":"Route Rurale Secondaire 1x2 voies", "devers":"G:3.0% | D:4.0%",
        "largeur":"2x3.0m voies + 1.0m accotement | dist axe/bord = 4.0m",
        "couches":"BB 5cm / GNT 20cm / Fond de Forme 25cm",
        "assain_g":"Caniveau macon 50x50 — H int=0.50m / dalle=10cm / radier=10cm / BP=20cm",
        "assain_d":"Caniveau macon 50x50 — H int=0.50m / dalle=10cm / radier=10cm / BP=20cm",
        "pente_long":f"{PL*100:.1f}%", "z0":Z0, "pk_fin":PK_FIN, "pas":PAS,
        "cv_g":"50x50 (H=0.50m)", "cv_d":"50x50 (H=0.50m)",
        "pente_g":0.030, "pente_d":0.040, "dist_g":4.0, "dist_d":4.0,
        "h_canal_g":0.50, "h_canal_d":0.50,
    }

# ═══════════════════════════════════════════════════════════════
#  DESSIN PROFIL EN TRAVERS
# ═══════════════════════════════════════════════════════════════
def draw_profil(meta):
    fig, ax = plt.subplots(figsize=(16, 6))
    ax.set_facecolor("#F8FAFC")
    fig.patch.set_facecolor("#0D1F38")

    pg=meta["pente_g"]; pd_=meta["pente_d"]
    dg=meta["dist_g"];  dd=meta["dist_d"]
    hg=meta["h_canal_g"]; hd=meta["h_canal_d"]
    z0=0.0
    z_bg = -dg * pg
    z_bd = -dd * pd_

    # Couches chaussée
    coul_ep  = [0.06, 0.10, 0.15, 0.20]
    coul_col = ["#2C2C2C","#555555","#888888","#B8936A"]
    coul_lab = ["BB Roulement","GB4 Base","GNT Sous-base","Fond de Forme"]
    cum = 0
    for ep, col, lab in zip(coul_ep, coul_col, coul_lab):
        px = [-dg, 0, dd, dd, 0, -dg]
        py = [z_bg-cum, z0-cum, z_bd-cum, z_bd-cum-ep, z0-cum-ep, z_bg-cum-ep]
        ax.fill(px, py, color=col, alpha=0.88, zorder=3)
        cum += ep

    # Surface chaussée
    ax.plot([-dg, 0, dd], [z_bg, z0, z_bd], color="#00FFFF", lw=2.5, zorder=6)
    ax.axvline(0, color="#FBBF24", lw=1.5, ls="--", zorder=7, alpha=0.9)
    ax.text(0, z0+0.08, "AXE", ha="center", color="#FBBF24", fontsize=9, fontweight="bold")

    # Dévers
    ax.annotate("", xy=(-dg, z_bg), xytext=(0, z0),
                arrowprops=dict(arrowstyle="->", color="#EF4444", lw=1.8))
    ax.text(-dg*0.55, (z0+z_bg)/2+0.15, f"G:{pg*100:.1f}%",
            color="#EF4444", fontsize=8, fontweight="bold", ha="center")
    ax.annotate("", xy=(dd, z_bd), xytext=(0, z0),
                arrowprops=dict(arrowstyle="->", color="#3B82F6", lw=1.8))
    ax.text(dd*0.55, (z0+z_bd)/2+0.15, f"D:{pd_*100:.1f}%",
            color="#3B82F6", fontsize=8, fontweight="bold", ha="center")

    # ── Caniveau GAUCHE ───────────────────────────────────────
    cw_g = min(hg + 0.20, 1.50)
    # Stratigraphie de bas en haut
    z_fdf_g = z_bg - 0.10 - hg - 0.20   # fond fouille
    z_bp_g  = z_fdf_g + 0.20             # dessus beton proprete
    z_rad_g = z_bp_g  + 0.10             # interieur fond canal
    z_tab_g = z_bg                        # tablier fini = bord chaussee
    cx_g = -dg - cw_g - 0.05

    # Fond fouille (hachures)
    ax.fill([cx_g-0.05, -dg+0.05, -dg+0.05, cx_g-0.05],
            [z_fdf_g, z_fdf_g, z_fdf_g-0.05, z_fdf_g-0.05],
            color="#D4B896", alpha=0.6, zorder=3)
    ax.text((cx_g + -dg)/2, z_fdf_g - 0.14, "Fond Fouille",
            ha="center", color="#92400E", fontsize=7, fontweight="bold")

    # Béton de propreté
    ax.fill([cx_g, -dg, -dg, cx_g],
            [z_fdf_g, z_fdf_g, z_bp_g, z_bp_g],
            color="#CBD5E1", alpha=0.8, zorder=4)
    ax.text((cx_g + -dg)/2, (z_fdf_g+z_bp_g)/2, "BP\n+20cm",
            ha="center", va="center", color="#334155", fontsize=6.5, fontweight="bold")

    # Radier canal
    ax.fill([cx_g, -dg, -dg, cx_g],
            [z_bp_g, z_bp_g, z_rad_g, z_rad_g],
            color="#9CA3AF", alpha=0.9, zorder=4)
    ax.text((cx_g + -dg)/2, (z_bp_g+z_rad_g)/2, f"Radier\n+10cm",
            ha="center", va="center", color="white", fontsize=6.5, fontweight="bold")

    # Corps canal (vide)
    ax.fill([cx_g, -dg, -dg, cx_g],
            [z_rad_g, z_rad_g, z_tab_g-0.10, z_tab_g-0.10],
            color="#E8F4FD", alpha=0.9, zorder=4, linewidth=0)
    ax.text((cx_g + -dg)/2, (z_rad_g + z_tab_g-0.10)/2,
            f"Canal\n{meta['cv_g']}", ha="center", va="center",
            color="#1E40AF", fontsize=7, fontweight="bold")

    # Dalle tablier
    ax.fill([cx_g-0.15, -dg+0.05, -dg+0.05, cx_g-0.15],
            [z_tab_g-0.10, z_tab_g-0.10, z_tab_g, z_tab_g],
            color="#6B2D8B", alpha=0.85, zorder=5)
    ax.text((cx_g-0.15 + -dg)/2, z_tab_g-0.05, "Tablier",
            ha="center", va="center", color="white", fontsize=6.5, fontweight="bold")

    # Contour caniveau G
    ax.plot([cx_g, cx_g, -dg, -dg, cx_g],
            [z_fdf_g-0.05, z_tab_g, z_tab_g, z_fdf_g-0.05, z_fdf_g-0.05],
            color="#9333EA", lw=2, zorder=6)
    ax.text(cx_g - 0.4, z_rad_g + hg/2, "GAUCHE",
            color="#EF4444", fontsize=9, fontweight="bold", ha="center", rotation=90)

    # ── Caniveau DROIT ────────────────────────────────────────
    cw_d = min(hd + 0.20, 1.20)
    z_fdf_d = z_bd - 0.10 - hd - 0.20
    z_bp_d  = z_fdf_d + 0.20
    z_rad_d = z_bp_d  + 0.10
    z_tab_d = z_bd
    cx_d_end = dd + cw_d + 0.05

    ax.fill([dd-0.05, cx_d_end+0.05, cx_d_end+0.05, dd-0.05],
            [z_fdf_d, z_fdf_d, z_fdf_d-0.05, z_fdf_d-0.05],
            color="#D4B896", alpha=0.6, zorder=3)
    ax.text((dd + cx_d_end)/2, z_fdf_d-0.14, "Fond Fouille",
            ha="center", color="#92400E", fontsize=7, fontweight="bold")

    ax.fill([dd, cx_d_end, cx_d_end, dd],
            [z_fdf_d, z_fdf_d, z_bp_d, z_bp_d],
            color="#CBD5E1", alpha=0.8, zorder=4)
    ax.text((dd+cx_d_end)/2, (z_fdf_d+z_bp_d)/2, "BP\n+20cm",
            ha="center", va="center", color="#334155", fontsize=6.5, fontweight="bold")

    ax.fill([dd, cx_d_end, cx_d_end, dd],
            [z_bp_d, z_bp_d, z_rad_d, z_rad_d],
            color="#9CA3AF", alpha=0.9, zorder=4)
    ax.text((dd+cx_d_end)/2, (z_bp_d+z_rad_d)/2, "Radier\n+10cm",
            ha="center", va="center", color="white", fontsize=6.5, fontweight="bold")

    ax.fill([dd, cx_d_end, cx_d_end, dd],
            [z_rad_d, z_rad_d, z_tab_d-0.10, z_tab_d-0.10],
            color="#FDF2F8", alpha=0.9, zorder=4)
    ax.text((dd+cx_d_end)/2, (z_rad_d+z_tab_d-0.10)/2,
            f"Canal\n{meta['cv_d']}", ha="center", va="center",
            color="#BE185D", fontsize=7, fontweight="bold")

    ax.fill([dd-0.05, cx_d_end+0.15, cx_d_end+0.15, dd-0.05],
            [z_tab_d-0.10, z_tab_d-0.10, z_tab_d, z_tab_d],
            color="#8B2D6B", alpha=0.85, zorder=5)
    ax.text((dd+cx_d_end)/2, z_tab_d-0.05, "Tablier",
            ha="center", va="center", color="white", fontsize=6.5, fontweight="bold")

    ax.plot([dd, dd, cx_d_end, cx_d_end, dd],
            [z_tab_d, z_fdf_d-0.05, z_fdf_d-0.05, z_tab_d, z_tab_d],
            color="#DB2777", lw=2, zorder=6)
    ax.text(cx_d_end+0.6, z_rad_d+hd/2, "DROIT",
            color="#3B82F6", fontsize=9, fontweight="bold", ha="center", rotation=90)

    # Cotes distances
    yref = min(z_bg, z_bd) - 1.6
    ax.annotate("", xy=(-dg, yref), xytext=(0, yref),
                arrowprops=dict(arrowstyle="<->", color="#94A3B8", lw=1.2))
    ax.text(-dg/2, yref-0.15, f"{dg:.2f} m", ha="center", color="#94A3B8", fontsize=8)
    ax.annotate("", xy=(dd, yref), xytext=(0, yref),
                arrowprops=dict(arrowstyle="<->", color="#94A3B8", lw=1.2))
    ax.text(dd/2, yref-0.15, f"{dd:.2f} m", ha="center", color="#94A3B8", fontsize=8)

    # Légende couches
    patchs = [mpatches.Patch(color=c, label=l) for c,l in zip(coul_col, coul_lab)]
    patchs += [
        mpatches.Patch(color="#9CA3AF", label="Radier canal (+10cm)"),
        mpatches.Patch(color="#CBD5E1", label="Beton proprete (+20cm)"),
        mpatches.Patch(color="#D4B896", label="Fond de fouille"),
    ]
    ax.legend(handles=patchs, loc="lower center", bbox_to_anchor=(0.5, -0.02),
              ncol=4, fontsize=7, facecolor="#1E3A5F", labelcolor="white",
              edgecolor="#00FFFF", framealpha=0.95)

    ax.set_title(f"Profil en travers — {meta['titre']}  |  PK0+000→PK{meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}  |  pente long. {meta['pente_long']}",
                 color="white", fontsize=10, fontweight="bold")
    ax.set_xlim(cx_g - 1.0, cx_d_end + 1.2)
    ax.set_ylim(min(z_fdf_g, z_fdf_d) - 0.5, z0 + 0.5)
    ax.set_xlabel("Distance / axe (m)", color="#94A3B8", fontsize=8)
    ax.set_ylabel("Cote relative (m)", color="#94A3B8", fontsize=8)
    ax.tick_params(colors="#94A3B8")
    for sp in ax.spines.values(): sp.set_edgecolor("#334155")
    ax.grid(True, alpha=0.2, color="#475569")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close()
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════════════
#  CONSTRUCTION DU CLASSEUR
# ═══════════════════════════════════════════════════════════════
def group_key(col):
    cu = col.upper()
    if cu == "PK":                     return "pk"
    if cu.startswith("AXE"):           return "axe"
    if cu.startswith("G_") and any(k in cu for k in
        ["BB","GB4","GNT","FOND","TRAV","BAU","ACC","VOIE","BORD_V","BORDURE","PST","ACCOT"]):
                                       return "trav_g"
    if cu.startswith("G_"):            return "ass_g"
    if cu.startswith("D_") and any(k in cu for k in
        ["BB","GB4","GNT","FOND","TRAV","BAU","ACC","VOIE","BORD_V","BORDURE","PST","ACCOT"]):
                                       return "trav_d"
    if cu.startswith("D_"):            return "ass_d"
    return "axe"

def build_wb(df_g, df_d, meta, path):
    wb = Workbook()

    def write_sheet(ws, df):
        ncols = len(df.columns)
        # LIGNE 1 = vrais noms colonnes (reconnus par RECEPTA)
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(1, ci, col)
            hdr, dat = PALETTE[group_key(col)]
            c.fill = fill(hdr)
            c.font = Font(color="FFFFFF", bold=True, size=8)
            c.alignment = CENTER
            c.border = BORDER
        ws.row_dimensions[1].height = 40

        # LIGNES 2+ = données
        for ri, row in df.iterrows():
            er = ri + 2
            for ci, (col, val) in enumerate(zip(df.columns, row), 1):
                c = ws.cell(er, ci, val)
                _, dat = PALETTE[group_key(col)]
                c.fill = fill(dat)
                c.font = Font(size=9, bold=(ci==1),
                              color="1E3A5F" if ci==1 else "111827")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = BORDER
            ws.row_dimensions[er].height = 14

        for i in range(1, ncols+1):
            ws.column_dimensions[get_column_letter(i)].width = 13
        ws.freeze_panes = "B2"

    ws_g = wb.active; ws_g.title = "Cote_Gauche"
    write_sheet(ws_g, df_g)

    ws_d = wb.create_sheet("Cote_Droit")
    write_sheet(ws_d, df_d)

    # LÉGENDE avec profil
    ws_l = wb.create_sheet("LEGENDE")
    ws_l.column_dimensions["A"].width = 4
    ws_l.column_dimensions["B"].width = 72

    img_buf = draw_profil(meta)
    img = XLImage(img_buf)
    img.width = 980; img.height = 360
    ws_l.add_image(img, "B2")
    for r in range(2, 25): ws_l.row_dimensions[r].height = 17

    strati = [
        "STRATIGRAPHIE ASSAINISSEMENT (de bas en haut) :",
        "  Fond de Fouille                         (cote la plus basse)",
        "  + 20 cm  → Beton de Proprete (BP)       dessus de la couche de 20cm",
        "  + 10 cm  → Radier Interieur             interieur fond du caniveau",
        "  + H canal (100 / 80 / 60 / 50 cm...)   hauteur utile interieure",
        "  + e dalle                               Tablier Fini = bord chaussee",
    ]
    specs = [(25, "SPECIFICATIONS DU PROJET", "0D1F38","00FFFF", True, 12)]
    for k, v in [
        ("Titre",         meta["titre"]),
        ("Devers",        meta["devers"]),
        ("Largeurs",      meta["largeur"]),
        ("Couches",       meta["couches"]),
        ("Pente long.",   meta["pente_long"]),
        ("Cote axe Z0",  f"{meta['z0']:.3f} m NGF"),
        ("PK debut/fin", f"0+000 / {meta['pk_fin']//1000}+{meta['pk_fin']%1000:03d}"),
        ("Pas PK",       f"{meta['pas']} m"),
    ]:
        specs.append((len(specs)+25, f"  {k+':':15s} {v}", "1E3A5F","FFFFFF", False, 10))

    specs.append((len(specs)+26, "", None, None, False, 6))
    specs.append((len(specs)+26, "ASSAINISSEMENT GAUCHE — "+meta["cv_g"], "6B2D8B","FFFFFF", True, 10))
    specs.append((len(specs)+26, "  "+meta["assain_g"], "6B2D8B","FFFFFF", False, 10))
    specs.append((len(specs)+26, "ASSAINISSEMENT DROIT  — "+meta["cv_d"], "8B2D6B","FFFFFF", True, 10))
    specs.append((len(specs)+26, "  "+meta["assain_d"], "8B2D6B","FFFFFF", False, 10))
    specs.append((len(specs)+26, "", None, None, False, 6))
    for s in strati:
        specs.append((len(specs)+26, s, "0F4C81","FFFFFF", s.startswith("STRAT"), 10))
    specs.append((len(specs)+26, "", None, None, False, 6))
    specs.append((len(specs)+26, "COLONNES COTE_GAUCHE", "1A6B3C","FFFFFF", True, 10))
    for col in df_g.columns:
        specs.append((len(specs)+26, f"  {col}", "1A6B3C","FFFFFF", False, 9))
    specs.append((len(specs)+26, "COLONNES COTE_DROIT", "155E2F","FFFFFF", True, 10))
    for col in df_d.columns:
        specs.append((len(specs)+26, f"  {col}", "155E2F","FFFFFF", False, 9))

    for row, txt, bg, fg, bold, sz in specs:
        c = ws_l.cell(row, 2, txt)
        if bg:
            c.fill = fill(bg)
            c.font = Font(color=fg, bold=bold, size=sz)
        else:
            c.font = Font(size=sz or 9)
        c.alignment = Alignment(vertical="center", indent=1)
        ws_l.row_dimensions[row].height = sz + 5 if sz else 8

    wb.save(path)
    print(f"OK  {path.split('/')[-1]:55s} {len(df_g):4d} PK x {len(df_g.columns)} cols")

# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    projets = [
        (projet_1, f"{OUT_DIR}/P1_RN_2x2voies_TPC_3km.xlsx"),
        (projet_2, f"{OUT_DIR}/P2_Urbain_Trottoirs_1.5km.xlsx"),
        (projet_3, f"{OUT_DIR}/P3_Autoroute_2x3_5km.xlsx"),
        (projet_4, f"{OUT_DIR}/P4_Rural_1x2_1km.xlsx"),
    ]
    print(f"{'Fichier':55s} {'PK':>5} {'Cols':>5}")
    print("-"*70)
    for fn, path in projets:
        df_g, df_d, meta = fn()
        build_wb(df_g, df_d, meta, path)

    print()
    print("Verification lecture RECEPTA (colonne PK + cotes numeriques) :")
    print("-"*70)
    for _, path in projets:
        df = pd.read_excel(path, sheet_name="Cote_Gauche", nrows=2)
        pk_col = next((c for c in df.columns if "PK" in str(c).upper()), None)
        num_cols = [c for c in df.columns
                    if c != pk_col and pd.api.types.is_numeric_dtype(df[c])]
        print(f"  {path.split('/')[-1]:50s}  PK='{pk_col}'  {len(num_cols)} cotes detectees")
    print()
    print(f"Tous les fichiers sont dans : {OUT_DIR}/")
